"""Python library extending yEd functionality through programmatic interface to graphs.

Currently supports...
* Building of yEd graph objects from scratch
* Reading of yEd graph files
* Management of yEd graph data in Spreadsheet (for simplified addition / management of data)
* Opening and basic control of yEd application

"""

from __future__ import annotations

# import asyncio
import io
import os
import platform
import re
import subprocess
import sys
import xml.etree.ElementTree as ET
from random import randint
from shutil import which
from time import sleep
from tkinter import messagebox as msg
from typing import Any, Dict, List, Optional, Union
from warnings import warn
from xml.dom import minidom

import openpyxl as pyxl
import pandas as pd
import psutil

# Enumerated parameters / Constants
app_platform = platform.platform().split("-")[0]
PROGRAM_NAME = "yEd"

# Testing related triggers
testing = False
local_testing = None
show_guis = True

LINE_TYPES = [
    "line",
    "dashed",
    "dotted",
    "dashed_dotted",
]

FONT_STYLES = [
    "plain",
    "bold",
    "italic",
    "bolditalic",
]

HORIZONTAL_ALIGNMENTS = [
    "left",
    "center",
    "right",
]

VERTICAL_ALIGNMENTS = [
    "top",
    "center",
    "bottom",
]

CUSTOM_PROPERTY_SCOPES = [
    "node",
    "edge",
]  # TODO: DOES THIS NEED GROUP?

CUSTOM_PROPERTY_TYPES = [
    "string",
    "int",
    "double",
    "boolean",
]


def checkValue(
    parameter_name: str,
    value: Any,
    validValues: Optional[List[str]] = None,
) -> None:
    """Check whether given inputs
    (e.g. Shape, Arrow type, Line type, etc.)
    are valid (within existing enumerated options).
    If not valid - returns error message for input."""

    if validValues:
        if value not in validValues:
            raise ValueError(f"{parameter_name} '{value}' is not supported. Use: '{', '.join(validValues)}'")


class File:
    """Object to check and act on (primarily yEd) files / filepaths."""

    def __init__(self, file_name_or_path=None, extension=""):
        self.DEFAULT_FILE_NAME = "temp"
        self.EXTENSION: str
        self.dir = self.path_validate(file_name_or_path)
        self.basename = self.base_name_validate(file_name_or_path, extension)
        self.fullpath = os.path.join(self.dir, self.basename)
        self.window_search_name = self.basename + " - yEd"
        self.file_exists = os.path.isfile(self.fullpath)

    def full_path_validate(self):
        self.file_exists = os.path.isfile(self.fullpath)

    def path_validate(self, temp_name_or_path=None):
        """Validate if the file was initialized with valid path - returning the same path - if not valid, return working directory as default path."""
        path = os.getcwd()
        if temp_name_or_path:
            path = os.path.dirname(temp_name_or_path)
            if not os.path.exists(path):
                path = os.getcwd()
        return os.path.realpath(path)

    def base_name_validate(self, temp_name_or_path, extension):
        """Validate / Build valid file name with extension"""
        temp_name = ""
        if temp_name_or_path:
            temp_name = os.path.basename(temp_name_or_path)

        # if no file name - give file name
        temp_name = temp_name or f"{self.DEFAULT_FILE_NAME}"

        # check for extension
        name, extension_from_name = os.path.splitext(temp_name)  # break up name into name + ext
        has_extension_assigned = any(extension_from_name)
        extension_received_from_function = any(extension)

        # if extension - update file object with Extension
        if has_extension_assigned:
            self.EXTENSION = extension_from_name

        # if no extension - assign extension from function and update file object
        elif not has_extension_assigned and extension_received_from_function:
            self.EXTENSION = extension
            temp_name += self.EXTENSION

        # if no extensions in name or from function - assign default file extension and update file object
        elif not has_extension_assigned and not extension_received_from_function:
            self.EXTENSION = ".graphml"  # default file extension
            temp_name += self.EXTENSION

        return temp_name  # give back name+ext combo

    def open_with_yed(self, force=False):
        """Method to open GraphML file directly with yEd application (must be installed and on path)."""

        # Ensure a valid yed extension file
        if self.EXTENSION != ".graphml":
            raise RuntimeWarning("Trying to open non-graphml file with yEd.")

        print("opening file with yed...")

        open_yed_file(self, force)

        return get_yed_pid()


class Label:
    """Generic Label Class for nodes / edges in yEd"""

    graphML_tagName = None

    def __init__(
        self,
        text="",
        height="18.1328125",
        width=None,
        alignment="center",
        font_family="Dialog",
        font_size="12",
        font_style="plain",
        horizontalTextPosition="center",
        underlined_text="false",
        text_color="#000000",
        icon_text_gap="4",
        horizontal_text_position="center",
        vertical_text_position="center",
        visible="true",
        border_color=None,
        background_color=None,
        has_background_color="false",
    ):
        # make class abstract
        if type(self) is Label:
            raise Exception("Label is an abstract class and cannot be instantiated directly")

        self._text = text

        # Initialize dictionary for parameters
        self._params = {}
        self.updateParam("horizontalTextPosition", horizontal_text_position, HORIZONTAL_ALIGNMENTS)
        self.updateParam("verticalTextPosition", vertical_text_position, VERTICAL_ALIGNMENTS)
        self.updateParam("alignment", alignment, HORIZONTAL_ALIGNMENTS)
        self.updateParam("fontStyle", font_style, FONT_STYLES)

        # TODO: Implement range checks
        self.updateParam("fontFamily", font_family)
        self.updateParam("iconTextGap", icon_text_gap)
        self.updateParam("fontSize", font_size)
        self.updateParam("textColor", text_color)
        self.updateParam("visible", visible.lower(), ["true", "false"])
        self.updateParam("underlinedText", underlined_text.lower(), ["true", "false"])
        if background_color:
            has_background_color = "true"
        self.updateParam("hasBackgroundColor", has_background_color.lower(), ["true", "false"])
        self.updateParam("width", width)
        self.updateParam("height", height)
        self.updateParam("borderColor", border_color)
        self.updateParam("backgroundColor", background_color)

    def updateParam(
        self,
        parameter_name,
        value,
        validValues=None,
    ):
        if value is None:
            return False
        checkValue(parameter_name, value, validValues)

        self._params[parameter_name] = value
        return True

    def addSubElement(self, shape):
        label = ET.SubElement(shape, self.graphML_tagName, **self._params)
        label.text = self._text


class NodeLabel(Label):
    """Node specific label"""

    VALIDMODELPARAMS = {
        "internal": ["t", "b", "c", "l", "r", "tl", "tr", "bl", "br"],
        "corners": ["nw", "ne", "sw", "se"],
        "sandwich": ["n", "s"],
        "sides": ["n", "e", "s", "w"],
        "eight_pos": ["n", "e", "s", "w", "nw", "ne", "sw", "se"],
    }

    graphML_tagName = "y:NodeLabel"

    def __init__(
        self,
        text,
        alignment="center",
        font_family="Dialog",
        font_size="12",
        font_style="plain",
        height="18.1328125",
        horizontalTextPosition="center",
        underlined_text="false",
        icon_text_gap="4",
        text_color="#000000",
        horizontal_text_position="center",
        vertical_text_position="center",
        visible="true",
        has_background_color="false",
        width="55.708984375",
        model_name="internal",
        border_color=None,
        background_color=None,
        model_position="c",
    ):
        super().__init__(
            text,
            height,
            width,
            alignment,
            font_family,
            font_size,
            font_style,
            horizontalTextPosition,
            underlined_text,
            text_color,
            icon_text_gap,
            horizontal_text_position,
            vertical_text_position,
            visible,
            border_color,
            background_color,
            has_background_color,
        )

        self.updateParam("modelName", model_name, NodeLabel.VALIDMODELPARAMS.keys())
        self.updateParam("modelPosition", model_position, NodeLabel.VALIDMODELPARAMS[model_name])


class EdgeLabel(Label):
    """Edge specific label"""

    VALIDMODELPARAMS = {
        "two_pos": ["head", "tail"],
        "centered": ["center"],
        "six_pos": ["shead", "thead", "head", "stail", "ttail", "tail"],
        "three_center": ["center", "scentr", "tcentr"],
        "center_slider": None,
        "side_slider": None,
    }

    graphML_tagName = "y:EdgeLabel"

    def __init__(
        self,
        text,
        alignment="center",
        font_family="Dialog",
        font_size="12",
        font_style="plain",
        height="18.1328125",
        horizontalTextPosition="center",
        underlined_text="false",
        icon_text_gap="4",
        text_color="#000000",
        horizontal_text_position="center",
        vertical_text_position="center",
        visible="true",
        has_background_color="false",
        width="55.708984375",
        model_name="centered",
        model_position="center",
        border_color=None,
        background_color=None,
        preferred_placement=None,
    ):
        super().__init__(
            text,
            height,
            width,
            alignment,
            font_family,
            font_size,
            font_style,
            horizontalTextPosition,
            underlined_text,
            text_color,
            icon_text_gap,
            horizontal_text_position,
            vertical_text_position,
            visible,
            border_color,
            background_color,
            has_background_color,
        )

        self.updateParam("modelName", model_name, EdgeLabel.VALIDMODELPARAMS.keys())
        self.updateParam("modelPosition", model_position, EdgeLabel.VALIDMODELPARAMS[model_name])
        self.updateParam("preferredPlacement", preferred_placement)


class CustomPropertyDefinition:
    """Custom properties which can be added to yEd objects / graph as a whole"""

    def __init__(
        self,
        scope,
        name,
        property_type,
        default_value,
    ):
        """
        scope: [node|edge]
        name: name of the custom property
        property_type: [string|boolean|int|double]
                        boolean: Java keywords [true|false]
        default_value: any above datatype represented as a string
        """
        self.scope = scope
        self.name = name
        self.property_type = property_type
        self.default_value = default_value
        self.id = "%s_%s" % (self.scope, self.name)

    def convert_to_xml(self) -> ET.Element:
        custom_prop_key = ET.Element("key", id=self.id)
        custom_prop_key.set("for", self.scope)
        custom_prop_key.set("attr.name", self.name)
        custom_prop_key.set("attr.type", self.property_type)

        return custom_prop_key


class Node:
    """yEd Node object - representing a single node in the graph"""

    custom_properties_defs = {}

    VALID_NODE_SHAPES = [
        "rectangle",
        "rectangle3d",
        "roundrectangle",
        "diamond",
        "ellipse",
        "fatarrow",
        "fatarrow2",
        "hexagon",
        "octagon",
        "parallelogram",
        "parallelogram2",
        "star5",
        "star6",
        "star6",
        "star8",
        "trapezoid",
        "trapezoid2",
        "triangle",
        "trapezoid2",
        "triangle",
    ]

    def __init__(
        self,
        name: str = "",  # non-unique node name
        label_alignment="center",
        shape="rectangle",
        font_family="Dialog",
        underlined_text="false",
        font_style="plain",
        font_size="12",
        shape_fill="#FFCC00",
        transparent="false",
        border_color="#000000",
        border_type="line",
        border_width="1.0",
        height=False,
        width=False,
        x=False,
        y=False,
        node_type="ShapeNode",
        UML: Union[bool, dict] = False,
        custom_properties=None,
        description="",
        url="",
    ):
        self.name: str = name
        self.id: str = generate_temp_uuid()  # temporary unique
        self.parent: Union[(Group, Graph, None)] = None

        self.list_of_labels: list[NodeLabel] = []  # initialize list of labels

        self.add_label(
            label_text=name,
            alignment=label_alignment,
            font_family=font_family,
            underlined_text=underlined_text,
            font_style=font_style,
            font_size=font_size,
        )

        self.node_type = node_type
        self.UML = UML

        # node shape
        checkValue("shape", shape, Node.VALID_NODE_SHAPES)
        self.shape = shape

        # shape fill
        self.shape_fill = shape_fill
        self.transparent = transparent

        # border options
        self.border_color = border_color
        self.border_width = border_width

        checkValue("border_type", border_type, LINE_TYPES)
        self.border_type = border_type

        # geometry
        self.geom = {}
        if height:
            self.geom["height"] = height
        if width:
            self.geom["width"] = width
        if x:
            self.geom["x"] = x
        if y:
            self.geom["y"] = y

        self.description = description
        self.url = url

        # Handle Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            if custom_properties:
                for k, v in custom_properties.items():
                    if k not in Node.custom_properties_defs:
                        raise RuntimeWarning("key %s not recognised" % k)
                    if name == k:
                        setattr(self, name, custom_properties[k])
                        break
                else:
                    setattr(self, name, definition.default_value)
            else:
                setattr(self, name, definition.default_value)

    def add_label(self, label_text, **kwargs) -> Node:
        """Adds node label - > returns node for continued node operations"""
        self.list_of_labels.append(NodeLabel(label_text, **kwargs))
        return self

    def convert_to_xml(self) -> ET.Element:
        """Converting node object to xml object"""

        xml_node = ET.Element("node", id=str(self.id))
        data = ET.SubElement(xml_node, "data", key="data_node")
        shape = ET.SubElement(data, "y:" + self.node_type)

        if self.geom:
            ET.SubElement(shape, "y:Geometry", **self.geom)
        # <y:Geometry height="30.0" width="30.0" x="475.0" y="727.0"/>

        ET.SubElement(shape, "y:Fill", color=self.shape_fill, transparent=self.transparent)

        ET.SubElement(
            shape,
            "y:BorderStyle",
            color=self.border_color,
            type=self.border_type,
            width=self.border_width,
        )

        for label in self.list_of_labels:
            label.addSubElement(shape)

        ET.SubElement(shape, "y:Shape", type=self.shape)

        # UML specific
        if self.UML:
            UML = ET.SubElement(shape, "y:UML")

            attributes = ET.SubElement(UML, "y:AttributeLabel", type=self.shape)
            attributes.text = self.UML["attributes"]

            methods = ET.SubElement(UML, "y:MethodLabel", type=self.shape)
            methods.text = self.UML["methods"]

            stereotype = self.UML["stereotype"] if "stereotype" in self.UML else ""
            UML.set("stereotype", stereotype)

        # Special items
        if self.url:
            url_node = ET.SubElement(xml_node, "data", key="url_node")
            url_node.text = self.url

        if self.description:
            description_node = ET.SubElement(xml_node, "data", key="description_node")
            description_node.text = self.description

        # Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            node_custom_prop = ET.SubElement(xml_node, "data", key=definition.id)
            node_custom_prop.text = getattr(self, name)

        return xml_node

    @classmethod
    def set_custom_properties_defs(cls, custom_property) -> None:
        cls.custom_properties_defs[custom_property.name] = custom_property


class Edge:
    """yEd Edge - connecting Nodes or Groups"""

    custom_properties_defs = {}

    ARROW_TYPES = [
        "none",
        "standard",
        "white_delta",
        "diamond",
        "white_diamond",
        "short",
        "plain",
        "concave",
        "convex",
        "circle",
        "transparent_circle",
        "dash",
        "skewed_dash",
        "t_shape",
        "crows_foot_one_mandatory",
        "crows_foot_many_mandatory",
        "crows_foot_many_optional",
        "crows_foot_one",
        "crows_foot_many",
        "crows_foot_optional",
    ]

    def __init__(
        self,
        node1: Node,
        node2: Node,
        name: str = "",
        arrowhead="standard",
        arrowfoot="none",
        color="#000000",
        line_type="line",
        width="1.0",
        label_background_color="",
        label_border_color="",
        source_label=None,
        target_label=None,
        custom_properties=None,
        description="",
        url="",
    ):
        # Primary operations
        self.node1: Node = node1
        self.node2: Node = node2
        self.name: str = name
        self.list_of_labels: list[EdgeLabel] = []  # initialize list of labels
        self.id: str = generate_temp_uuid()  # give temp id
        self.parent: Union[(Group, Graph, None)] = None

        if name:
            self.add_label(
                name,
                border_color=label_border_color,
                background_color=label_background_color,
            )

        # if not node1 or not node2:
        #     id = "%s_%s" % (node1, node2)

        if source_label is not None:
            self.add_label(
                source_label,
                model_name="six_pos",
                model_position="shead",
                preferred_placement="source_on_edge",
                border_color=label_border_color,
                background_color=label_background_color,
            )

        if target_label is not None:
            self.add_label(
                target_label,
                model_name="six_pos",
                model_position="shead",
                preferred_placement="source_on_edge",
                border_color=label_border_color,
                background_color=label_background_color,
            )

        checkValue("arrowhead", arrowhead, Edge.ARROW_TYPES)
        self.arrowhead = arrowhead

        checkValue("arrowfoot", arrowfoot, Edge.ARROW_TYPES)
        self.arrowfoot = arrowfoot

        checkValue("line_type", line_type, LINE_TYPES)
        self.line_type = line_type

        self.color = color
        self.width = width

        self.description = description
        self.url = url

        # Handle Edge Custom Properties
        for name, definition in Edge.custom_properties_defs.items():
            if custom_properties:
                for k, v in custom_properties.items():
                    if k not in Edge.custom_properties_defs:
                        raise RuntimeWarning("key %s not recognised" % k)
                    if name == k:
                        setattr(self, name, custom_properties[k])
                        break
                else:
                    setattr(self, name, definition.default_value)
            else:
                setattr(self, name, definition.default_value)

    def add_label(self, label_text, **kwargs):
        """Adding edge label"""
        self.list_of_labels.append(EdgeLabel(label_text, **kwargs))
        # Enable method chaining
        return self

    def convert_to_xml(self) -> ET.Element:
        """Converting edge object to xml object"""

        edge = ET.Element(
            "edge",
            id=str(self.id),
            source=str(self.node1.id),
            target=str(self.node2.id),
        )

        data = ET.SubElement(edge, "data", key="data_edge")
        pl = ET.SubElement(data, "y:PolyLineEdge")

        ET.SubElement(pl, "y:Arrows", source=self.arrowfoot, target=self.arrowhead)
        ET.SubElement(pl, "y:LineStyle", color=self.color, type=self.line_type, width=self.width)

        for label in self.list_of_labels:
            label.addSubElement(pl)

        if self.url:
            url_edge = ET.SubElement(edge, "data", key="url_edge")
            url_edge.text = self.url

        if self.description:
            description_edge = ET.SubElement(edge, "data", key="description_edge")
            description_edge.text = self.description

        # Edge Custom Properties
        for name, definition in Edge.custom_properties_defs.items():
            edge_custom_prop = ET.SubElement(edge, "data", key=definition.id)
            edge_custom_prop.text = getattr(self, name)

        return edge

    @classmethod
    def set_custom_properties_defs(cls, custom_property) -> None:
        cls.custom_properties_defs[custom_property.name] = custom_property


class Group:
    """yEd Group Object (Visual Container of Nodes / Edges / also can recursively act as Node)"""

    VALID_SHAPES = [
        "rectangle",
        "rectangle3d",
        "roundrectangle",
        "diamond",
        "ellipse",
        "fatarrow",
        "fatarrow2",
        "hexagon",
        "octagon",
        "parallelogram",
        "parallelogram2",
        "star5",
        "star6",
        "star6",
        "star8",
        "trapezoid",
        "trapezoid2",
        "triangle",
        "trapezoid2",
        "triangle",
    ]

    def __init__(
        self,
        name: str = "",
        top_level_graph=None,
        label_alignment="center",
        shape="rectangle",
        closed="false",
        font_family="Dialog",
        underlined_text="false",
        font_style="plain",
        font_size="12",
        fill="#FFCC00",
        transparent="false",
        border_color="#000000",
        border_type="line",
        border_width="1.0",
        height=False,
        width=False,
        x=False,
        y=False,
        custom_properties=None,
        description="",
        url="",
    ):
        self.name = name
        self.parent: Union[(Group, Graph, None)] = None  # set during add_group
        self.id = generate_temp_uuid()

        self.nodes: dict[str, Node] = {}
        self.groups: dict[str, Group] = {}
        self.edges: dict[str, Edge] = {}
        self.combined_objects = {}

        self.top_level_graph = top_level_graph

        # node shape
        checkValue("shape", shape, Group.VALID_SHAPES)
        self.shape = shape

        self.closed = closed

        # label formatting options
        self.font_family = font_family
        self.underlined_text = underlined_text

        checkValue("font_style", font_style, FONT_STYLES)
        self.font_style = font_style
        self.font_size = font_size

        checkValue("label_alignment", label_alignment, HORIZONTAL_ALIGNMENTS)
        self.label_alignment = label_alignment

        self.fill = fill
        self.transparent = transparent

        self.geom = {}
        if height:
            self.geom["height"] = height
        if width:
            self.geom["width"] = width
        if x:
            self.geom["x"] = x
        if y:
            self.geom["y"] = y

        self.border_color = border_color
        self.border_width = border_width

        checkValue("border_type", border_type, LINE_TYPES)
        self.border_type = border_type

        self.description = description
        self.url = url

        # Handle Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            if custom_properties:
                for k, v in custom_properties.items():
                    if k not in Node.custom_properties_defs:
                        raise RuntimeWarning("key %s not recognised" % k)
                    if name == k:
                        setattr(self, name, custom_properties[k])
                        break
                else:
                    setattr(self, name, definition.default_value)
            else:
                setattr(self, name, definition.default_value)

    def add_node(self, node: Union[Node, str, None] = None, **kwargs) -> Node:
        """Adding node within Group - accepts node object (simply assigns), or node name or none (to create new node without name)."""
        return add_node(self, node, **kwargs)

    def add_group(self, group: Union[Group, str, None] = None, **kwargs) -> Group:
        """Adding group to Group - accepts group object (simply assigns), or group name or none (to create new group without name)"""
        return add_group(self, group, **kwargs)

    def add_edge(
        self,  # owner
        node1: Optional[Union[(Node, Group, str)]] = None,
        node2: Optional[Union[(Node, Group, str)]] = None,
        **kwargs,
    ) -> Edge:
        """Adding edge to Group - for node1/node2 uses node / group objects or accepts names (creating new nodes under self) - or function can alternatively accept an instantiated Edge object."""

        # map args into kwargs in case of spreadsheet data management ops
        if node1:
            kwargs["node1"] = node1
        if node2:
            kwargs["node2"] = node2

        return add_edge(self, **kwargs)

    # Removal of items ==============================
    def remove_node(self, node: Union[Node, str]) -> None:
        """Remove/Delete a node from group - by object or id."""
        remove_node(self, node)

    def remove_group(self, group: Union[Group, str], **kwargs) -> None:
        """Removes a group from within current group object (and same parent graph) - by object or id."""
        remove_group(self, group, **kwargs)

    def remove_edge(self, edge: Union[Edge, str]) -> None:
        """Removing edge from group - by object or id."""
        remove_edge(self, edge)

    def is_ancestor(self, node) -> bool:
        """Check for possible nesting conflict of this id usage"""
        return node.parent is not None and (node.parent is self or self.is_ancestor(node.parent))

    def convert_to_xml(self) -> ET.Element:
        """Converting graph object to graphml xml object"""

        node = ET.Element("node", id=self.id)
        node.set("yfiles.foldertype", "group")
        data = ET.SubElement(node, "data", key="data_node")

        # node for group
        pabn = ET.SubElement(data, "y:ProxyAutoBoundsNode")
        r = ET.SubElement(pabn, "y:Realizers", active="0")
        group_node = ET.SubElement(r, "y:GroupNode")

        if self.geom:
            ET.SubElement(group_node, "y:Geometry", **self.geom)

        ET.SubElement(group_node, "y:Fill", color=self.fill, transparent=self.transparent)

        ET.SubElement(
            group_node,
            "y:BorderStyle",
            color=self.border_color,
            type=self.border_type,
            width=self.border_width,
        )

        label = ET.SubElement(
            group_node,
            "y:NodeLabel",
            modelName="internal",
            modelPosition="t",
            fontFamily=self.font_family,
            fontSize=self.font_size,
            underlinedText=self.underlined_text,
            fontStyle=self.font_style,
            alignment=self.label_alignment,
        )
        label.text = self.name

        ET.SubElement(group_node, "y:Shape", type=self.shape)

        ET.SubElement(group_node, "y:State", closed=self.closed)

        graph = ET.SubElement(node, "graph", edgedefault="directed", id=self.id)

        if self.url:
            url_node = ET.SubElement(node, "data", key="url_node")
            url_node.text = self.url

        if self.description:
            description_node = ET.SubElement(node, "data", key="description_node")
            description_node.text = self.description

        # Add group contained items (recursive)
        for id in self.nodes:
            n = self.nodes[id].convert_to_xml()
            graph.append(n)

        for id in self.groups:
            n = self.groups[id].convert_to_xml()
            graph.append(n)

        for id in self.edges:
            e = self.edges[id].convert_to_xml()
            graph.append(e)

        # Node Custom Properties
        for name, definition in Node.custom_properties_defs.items():
            node_custom_prop = ET.SubElement(node, "data", key=definition.id)
            node_custom_prop.text = getattr(self, name)

        return node
        # ProxyAutoBoundsNode crap just draws bar at top of group


class GraphStats:
    """Object to query and carry complete structure of current (recursive) graph objects and relationships."""

    def __init__(self, graph: Graph):
        self.graph = graph
        self.gather_metadata()  # initial extraction

    def recursive_id_extract(self, graph_or_input_node) -> None:
        """Gather complete structure of current (recursive) graph objects and relationships."""
        sub_nodes = graph_or_input_node.nodes.values()
        sub_groups = graph_or_input_node.groups.values()
        sub_edges = graph_or_input_node.edges.values()

        for node in sub_nodes:
            self.all_nodes[node.id] = node

        for edge in sub_edges:
            self.all_edges[edge.id] = edge

        for group in sub_groups:
            self.all_groups[group.id] = group
            self.recursive_id_extract(group)

    def gather_metadata(self):
        """Gather metadata for all objects in the graph."""

        # Establish / clear data structures
        self.all_nodes: dict[str, Node] = {}
        self.all_groups: dict[str, Group] = {}
        self.all_objects: dict[str, Union[Node, Group]] = {}
        self.all_edges: dict[str, Edge] = {}
        self.all_graph_items: dict[str, Union[Node, Group, Edge]] = {}
        self.id_to_name: dict[str, str] = {}
        self.name_to_ids: dict[str, list[str]] = {}
        self.duplicate_names: set[str] = set()

        # (re)extract core graph data
        self.recursive_id_extract(self.graph)

        # Combine remaining data ========================
        self.all_objects = {**self.all_nodes, **self.all_groups}
        self.all_graph_items = {**self.all_objects, **self.all_edges}
        for obj in self.all_graph_items.values():
            self.id_to_name[obj.id] = obj.name
            current_ids = self.name_to_ids.get(obj.name, [])
            if current_ids:
                self.duplicate_names.add(obj.name)
            current_ids.append(obj.id)
            self.name_to_ids[obj.name] = current_ids

    def find_by_id(self, id) -> Union[Node, Group, Edge, None]:
        """Find object by unique yEd id."""
        return self.all_graph_items.get(id, None)

    def find_by_name(self, name) -> List[Union[Node, Group, Edge]]:
        """Find object by user assigned name - needs to provide for multiple (needs deduplication)."""
        return [self.all_graph_items[id] for id in self.name_to_ids.get(name, [])]

    def name_reused(self, name: str) -> bool:
        """Find object by user assigned name - needs to provide for multiple (needs deduplication)."""
        return name in self.duplicate_names

    def print_stats(self):
        print(f"Graph Stats of {self.graph.id}")
        for k, v in vars(self).items():
            if isinstance(v, Graph):
                continue
            else:
                print(f"\tStat: {k} : {len(v)}")


class SpreadsheetManager:
    """Object to handle interfacing of graphs with Spreadsheet in bulk data management operations."""

    def __init__(self):
        # Template operations ==============
        self.WB_TYPES = ["obj_and_hierarchy", "object_data", "relations"]
        self.TEMP_XLSX_WORKBOOK = File("yedextended_temp.xlsx").fullpath
        self.OBJECTS_WS_NAME = "Objects_and_Groups"
        self.RELATIONS_WS_NAME = "Relations"
        self.DISAMBIGUATING_SEPARATOR = "##ID:##"

        # Graph operations ================
        self.graph = Graph()
        self.original_stats: GraphStats
        # self.original_id_to_label: dict[str, str] = dict()
        # self.original_id_to_obj: dict[str, Union[(Group, Edge, Node)]] = dict()
        # self.obj_keys = list()
        # self.obj_values = list()
        # self.dup_objects = list()

    def kill_spreadsheet(self) -> None:
        """Providing a utility to help primarily during test"""
        if app_platform == "Windows":
            os.system('taskkill /f /im "excel.exe"')  # FIXME: should support libreoffice, etc
        elif app_platform == "Linux":
            command = "ps aux | grep libreoffice | grep -v grep | awk '{print $2}' | xargs kill"
            return execute_shell_command(command)
            # raise NotImplementedError(
            #     "Spreadsheet app - killing functionality not yet implemented."
            # )
            # command = [PROGRAM_NAME, file.fullpath]
            # command = f"{PROGRAM_NAME} '{file.fullpath}'"
            # output = execute_shell_command(command)
            # subprocess.Popen(command, shell=True, start_new_session=True)

    def bulk_data_op_verify(self, type) -> None:
        """Check if the given bulk data management type is valid for spreadsheet operations."""
        self.type = type or self.WB_TYPES[0]  # default
        if self.type not in self.WB_TYPES:
            raise RuntimeWarning("Invalid spreadsheet type. Use: %s" % ", ".join(self.WB_TYPES))

    def create_spreadsheet_template(self, type) -> None:
        """Generate spreadsheet wb per template for that wb type."""

        self.bulk_data_op_verify(type)

        if os.path.isfile(self.TEMP_XLSX_WORKBOOK):
            os.remove(self.TEMP_XLSX_WORKBOOK)

        # create workbook
        spreadsheet_wb = pyxl.Workbook()

        # Inserting /organizing sheets
        spreadsheet_ws = spreadsheet_wb.active

        # Create object ws
        objects_ws = spreadsheet_wb.create_sheet(self.OBJECTS_WS_NAME, 0)

        # Add header to sheet
        objects_ws.cell(
            row=1,
            column=1,
            value="FORMAT -> OBJECT_LABEL | OBJECT_ID (OPTIONAL TO SUPPORT DISAMBIGUATION) - NOTE: INDENTATION OF INFO ONE COLUMN DESIGNATES BELONGING TO OBJECT ABOVE.",
        )

        # Add relations sheet and header
        if self.type == "relations":
            relations_ws = spreadsheet_wb.create_sheet(self.RELATIONS_WS_NAME, 1)
            relations_ws.cell(
                row=1,
                column=1,
                value="FORMAT -> NODE1_LABEL | NODE2_LABEL | EDGE_LABEL (OPTIONAL) | EDGE_OWNER (OPTIONAL - TO ASSIGN OWNERSHIP TO SPECIFIC GROUP) - NOTE: DISAMBIGUATION SUPPORTED BY CONCATENATING LABEL WITH '##ID:##'+ID SHARED WITH OBJECT",
            )

        # Clean up
        if spreadsheet_ws:
            spreadsheet_wb.remove(spreadsheet_ws)  # removing default sheet
        self.kill_spreadsheet()
        spreadsheet_wb.save(self.TEMP_XLSX_WORKBOOK)
        spreadsheet_wb.close()

    def open_close_spreadsheet_app(*args, **kwargs):
        """Decorator to handle opening, saving, and closing spreadsheet operations."""
        save = kwargs.get("save", False)

        def decorator(func):
            def wrapper_func(self, *args, **kwargs):
                in_mem_file = None
                spreadsheet_data = kwargs.get("spreadsheet_data", None)
                type = kwargs.get("type", None)
                self.bulk_data_op_verify(type)

                # Graph to spreadsheet
                if save:
                    self.create_spreadsheet_template(type)
                    sleep(0.5)  # FIXME: Unclear if necessary
                    with open(self.TEMP_XLSX_WORKBOOK, "rb") as f:
                        in_mem_file = io.BytesIO(f.read())

                # Spreadsheet to graph
                else:
                    # In case we are given a file path or in-memory file, for spreadsheet to graph
                    if spreadsheet_data:
                        if isinstance(spreadsheet_data, io.BytesIO):
                            in_mem_file = spreadsheet_data
                        elif isinstance(spreadsheet_data, str):
                            with open(spreadsheet_data, "rb") as f:
                                in_mem_file = io.BytesIO(f.read())
                    # Primary use case - from template for graph to spreadsheet
                    else:
                        # It is assumed that the template is already created and filled during graph_to_spreadsheet at this point
                        # Lets pull a version into memory
                        with open(self.TEMP_XLSX_WORKBOOK, "rb") as f:
                            in_mem_file = io.BytesIO(f.read())

                # If nothing in memory at this point we have an issue
                if not in_mem_file:
                    raise RuntimeWarning("No spreadsheet data found to open.")

                # provide fresh handles
                self.spreadsheet_wb = pyxl.load_workbook(in_mem_file)
                self.objects_ws = self.spreadsheet_wb[self.OBJECTS_WS_NAME]
                if self.type == "relations":
                    self.relations_ws = self.spreadsheet_wb[self.RELATIONS_WS_NAME]

                # Perform functions operations
                func(self, *args, **kwargs)

                # Clean up
                if save:
                    self.spreadsheet_wb.save(self.TEMP_XLSX_WORKBOOK)
                self.kill_spreadsheet()

            return wrapper_func

        return decorator

    def disambiguate_object(self, obj: Union[Node, Group, str], direction: str = "out"):
        """Generate Name:ID string for object - disambiguate if needed.
        direction: [in|out] - in for spreadsheet to graph, out for graph to spreadsheet
        """

        if direction == "out":
            if not obj:
                return None
            if obj.name in self.original_stats.duplicate_names:  # migrate to using graphstats
                return obj.name + self.DISAMBIGUATING_SEPARATOR + obj.id  # FIXME: IN SPREADSHEET_TO_GRAPH
            else:
                return obj.name
        elif direction == "in":
            if not obj:
                return None, None

            if self.DISAMBIGUATING_SEPARATOR in obj:
                name = obj.split(self.DISAMBIGUATING_SEPARATOR)[0]
                id = obj.split(self.DISAMBIGUATING_SEPARATOR)[1]
                if not all([name, id]):
                    warn(
                        f"Invalid deduplication format of edge information or empty name: {name}:{id}.",
                        SyntaxWarning,
                    )
                    return name, id
                # Normal situation in deduplication notation
                else:
                    return name, id
            # For normal non duplicate case
            else:
                return obj, None  # name,id

    @open_close_spreadsheet_app(save=True)
    def graph_to_spreadsheet_conversion(self, type=None, graph=None) -> None:
        """Converting graph object to spreadsheet sheet format for bulk data management operations."""
        # lets assume there can be multiple graphs in session - so we should pass
        if graph:
            self.graph = graph

        # Lets gather starting stats
        self.original_stats = self.graph.gather_graph_stats()

        def graph_object_extract_to_spreadsheet(self, input_node: Union[Group, Graph], indent_level):
            """Extracting graph objects to spreadsheet."""
            nonlocal row

            sub_nodes = input_node.nodes
            sub_groups = input_node.groups

            no_sub_items = not any([sub_nodes, sub_groups])

            for node in sub_nodes.values():
                # desc = node.get(description, "")
                # url = node.get(url, "")

                # posting to spreadsheet
                self.objects_ws.cell(row=row, column=indent_level, value=str(node.name))
                self.objects_ws.cell(row=row, column=indent_level + 1, value=str(node.id))
                row += 1

            for group in sub_groups.values():
                # id = group.id or ""
                # label = getattr(group, "label", "")

                # posting to spreadsheet
                self.objects_ws.cell(row=row, column=indent_level, value=str(group.name))
                self.objects_ws.cell(row=row, column=indent_level + 1, value=str(group.id))
                row += 1

                # Recursive extraction - adapting indent
                graph_object_extract_to_spreadsheet(self, group, indent_level=indent_level + 1)

            # Hacky workaround for tabulated ownership structure
            #     Add empty node under group
            if isinstance(input_node, Group) and no_sub_items:
                self.objects_ws.cell(row=row, column=indent_level, value="empty_group_Node")
                row += 1

        def relations_extract_to_spreadsheet(self, input_node: Union[Group, Graph]):
            """Extract relations recursively - providing owner if needed"""
            nonlocal row

            # Go through edges of this "owning" object
            sub_edges = input_node.edges
            for edge in sub_edges.values():
                node1name = self.disambiguate_object(edge.node1)
                node2name = self.disambiguate_object(edge.node2)

                group_name = ""
                if isinstance(input_node, Group):
                    group_name = self.disambiguate_object(input_node)

                edge_name = self.disambiguate_object(edge)

                # post to spreadsheet
                self.relations_ws.cell(row=row, column=col, value=node1name)
                self.relations_ws.cell(row=row, column=col + 1, value=node2name)
                self.relations_ws.cell(row=row, column=col + 2, value=edge_name)
                self.relations_ws.cell(row=row, column=col + 3, value=group_name)
                row += 1

            # Go to next level of relations / ownership - recursive
            sub_groups = input_node.groups
            for group in sub_groups.values():
                relations_extract_to_spreadsheet(self, group)

        # Perform the transformation to spreadsheet ========================
        if self.type == "obj_and_hierarchy" or self.type == "relations":
            row = 2
            graph_object_extract_to_spreadsheet(self, self.graph, indent_level=1)

        if self.type == "relations":
            row = 2
            col = 1
            relations_extract_to_spreadsheet(self, self.graph)

    @open_close_spreadsheet_app(save=False)
    def spreadsheet_to_graph_conversion(self, type: Optional[str] = None, spreadsheet_data=None):
        """Converting spreadsheet sheet data back into graph object."""
        self.bulk_data_op_verify(type)

        # Update original stats
        self.original_stats = self.graph.gather_graph_stats()

        # Begin transformation from spreadsheet to graph ========================
        if self.type == "obj_and_hierarchy":
            # Pull out object data
            obj_data = list(self.objects_ws.values)
            obj_data.pop(0)  # remove header

            # identifying indents in spreadsheet (marker for groupings)
            indent: list[int] = []
            for row_count, row in enumerate(obj_data):
                number_empty_cells = 0
                if row_count == 0:
                    pass  # just log 0
                else:
                    for val in row:
                        # if the cell is empty - increase count
                        if val == None or val == "":
                            number_empty_cells += 1
                        else:
                            break  # done incrementing - time to record
                indent.append(number_empty_cells)

            # identifying groups
            num_items = len(obj_data)
            group_identifiers = list(
                map(
                    lambda x: 1 if indent[x + 1] > indent[x] else 0,
                    range(0, num_items - 1),
                )
            )  # FIXME: WOULD FAIL ON A GROUP WITHOUT NAME
            group_identifiers.append(0)  # small limitation - deepest or last cannot be group - must have submembers

            # sorting ownership based on indents/groups - also correcting indents if name = ""
            owner_indexing: dict[int, Union[int, None]] = dict()
            indent_and_group_ident = zip(indent, group_identifiers)
            indent_and_group_ident = [list(row) for row in indent_and_group_ident]
            for i in range(0, num_items):
                owner_indexing[i] = None  # default value - no object owner - graph owner

                # first item has no object owner - graph owner
                if i == 0:
                    continue

                curr_indent = indent[i]
                for j, (indent_i, is_group) in enumerate(reversed(indent_and_group_ident[:i])):
                    # if indent_i == curr_indent - 1 and is_group == 1: # going only on 1 tab indent (no support for empty name)
                    if (
                        indent_i < curr_indent and is_group == 1
                    ):  # just going based on less indent (special case - empty name leads to extra indent)
                        owner_indexing[i] = i - (j + 1)
                        indent_and_group_ident[i][0] = indent_i + 1
                        indent[i] = indent_i + 1
                        break

            # Building / Modifying objects
            objects = list()
            object_id_mappings: dict[str, str] = dict()
            id_replaced_by_other_mapping: dict[str, Union[(Node, Group)]] = dict()
            # all_bulk_mod_ids = set()
            # all_curr_obj_ids = set(self.original_id_to_obj.keys())
            for i, (starting_indent, gr_i, obj_row) in enumerate(zip(indent, group_identifiers, obj_data)):
                # Extracting label and id
                name = str(
                    obj_row[starting_indent] or ""
                )  # assumes name is given and we need to convert to string for some items
                id = None
                try:
                    id = obj_row[starting_indent + 1]
                except Exception as e:
                    print(f"Node missing Id: {e}.")

                # Checks
                id_exists = id is not None
                id_found = id in self.original_stats.all_objects.keys()

                is_group = gr_i == 1
                is_node = gr_i == 0
                is_group_owned = owner_indexing[i] is not None
                if is_group_owned:
                    owner_index = owner_indexing[i]  # this works because the owner is always before
                    owner_object = objects[owner_index]
                else:
                    owner_object = self.graph

                # Early exit for group identifying nodes (hacky workaround)
                if name == "empty_group_Node":
                    objects.append(None)
                    continue

                # This is a new object - create it
                if not id_found:
                    # Add group
                    if is_group:
                        new_group = owner_object.add_group(name)
                        objects.append(new_group)

                        if id_exists:
                            object_id_mappings[id] = new_group.id

                    # Add node
                    elif is_node:
                        new_node = owner_object.add_node(name)
                        objects.append(new_node)
                        if id_exists:
                            object_id_mappings[id] = new_node.id
                    else:
                        raise NotImplementedError("This state not implemented")

                elif id_exists:  # TODO: NEED TO FINISH THIS
                    existing_obj = self.original_stats.all_objects[id]  # FIXME:

                    # change from node -> group
                    if isinstance(existing_obj, Node) and is_group:
                        # remove node
                        try:
                            existing_obj.parent.remove_node(existing_obj)
                        except Exception:
                            warn("Node no longer existing - to remove")

                        # add group
                        new_group = owner_object.add_group(name)
                        objects.append(new_group)
                        object_id_mappings[id] = new_group.id

                        # TODO: ADD CONVERSION MAPPING - ID CHANGES

                    # changed structuring from node <-> group
                    # group -> node
                    elif isinstance(existing_obj, Group) and not is_group:
                        # remove group (without changing dependencies)
                        try:
                            existing_obj.parent.remove_group(existing_obj, heal=False)
                        except Exception:
                            warn("Group no longer existing - to remove")

                        # add node
                        new_node = owner_object.add_node(name)
                        objects.append(new_node)
                        object_id_mappings[id] = new_node.id

                    # ownership changed - update
                    elif existing_obj.parent is not owner_object:
                        if is_group:
                            try:
                                existing_obj.parent.remove_group(existing_obj, heal=False)
                            except Exception:
                                warn("Group no longer existing - to remove")

                            mod_group = owner_object.add_group(name)
                            objects.append(mod_group)
                            object_id_mappings[id] = mod_group.id

                        elif is_node:
                            try:
                                existing_obj.parent.remove_node(existing_obj)
                            except Exception:
                                warn("Node no longer existing - to remove")

                            mod_node = owner_object.add_node(name)
                            objects.append(mod_node)
                            object_id_mappings[id] = mod_node.id

                    # no changes to object (so wont need to be deleted)
                    else:
                        objects.append(existing_obj)

                    # just name changed
                    if existing_obj.name != name:
                        existing_obj.name = name

            # Deleted objects - items previously with ids and ids are no longer there
            # Finding difference of ids - previous ids no longer there... #FIXME: WHAT ABOUT CHANGED IDS?
            all_updated_obj = objects.copy()
            all_orig_obj = list(self.original_stats.all_objects.values())
            all_deleted_obj = [obj for obj in all_orig_obj if obj not in all_updated_obj and obj is not None]
            for obj in all_deleted_obj:
                parent = obj.parent or self.graph
                if isinstance(obj, Group):  # group
                    # find all immediate dependents and connect them to owner
                    try:
                        parent.remove_group(obj, heal=False)
                    except Exception:
                        warn(f"Group {obj.name} no longer existing - to remove")
                elif isinstance(obj, Node):  # node
                    try:
                        parent.remove_node(obj)
                    except Exception:
                        warn(f"Node {obj.name} no longer existing - to remove")

            # Update all ids - after delete operations
            for obj in objects:
                if obj:
                    assign_traceable_id(obj)

        elif self.type == "relations":
            # Access the relations sheet
            self.relations_ws = self.spreadsheet_wb[self.RELATIONS_WS_NAME]
            relations_data = self.relations_ws.values
            row_length = None

            # declarations
            edge_ids_after_mod = set()

            # process sheet rows
            count: int = 0
            for row in relations_data:
                if count == 0:
                    row_length = len(row)
                    count += 1
                    continue

                # Declarations ===================================
                # names
                node1_name: str = ""
                node2_name: str = ""
                edge_name: str = ""
                owner_name: str = ""

                # ids
                node1_id: str = ""
                node2_id: str = ""
                edge_id: str = ""
                owner_id: str = ""

                # found flag
                node1_found = False
                node2_found = False
                edge_found = False
                owner_found = False

                # separating row values into variables (before disambiguation)
                if row_length == 4:
                    node1_name, node2_name, edge_name, owner_name = row
                    edge_id = str(edge_id)
                elif row_length == 3:
                    node1_name, node2_name, edge_name = row
                elif row_length == 2:
                    node1_name, node2_name = row

                # Quick check if we are working with minimally complete data
                # At this point the name is name(+id) - assumed we can ignore empty nodes / warn
                two_node_check = node1_name is not None and node2_name is not None
                if not two_node_check:
                    warn(
                        "Node names not found...skipping line of Relations.",
                        SyntaxWarning,
                    )
                    continue

                # Disambiguate the object information based on constant separator
                node1_name, node1_id = self.disambiguate_object(node1_name, direction="in")
                node2_name, node2_id = self.disambiguate_object(node2_name, direction="in")
                edge_name, edge_id = self.disambiguate_object(edge_name, direction="in")
                owner_name, owner_id = self.disambiguate_object(owner_name, direction="in")

                # Try to reidentify items ========================================
                def find_checks(name, id):
                    result_object = None
                    id_found = False
                    name_found = False

                    # look for ID - in case of disambiguation
                    if id:
                        id_found = id in self.original_stats.all_graph_items.keys()
                        if id_found:
                            result_object = self.original_stats.all_graph_items[id]

                    if not result_object:
                        # Look by name
                        if name:
                            name_found = (
                                name in self.original_stats.name_to_ids.keys()
                                and not self.original_stats.name_reused(name)
                            )
                            if name_found:
                                result_object = self.original_stats.all_graph_items[
                                    self.original_stats.name_to_ids[name][0]
                                ]

                    return (
                        result_object,
                        result_object is not None,
                    )  # , any(id_found, name_found), id_found, name_found

                # Looking for edge =================================
                edge_object, edge_found = find_checks(edge_name, edge_id)

                # Looking for node 1 ================================
                node1_object, node1_found = find_checks(node1_name, node1_id)

                # Looking for node 2 =================================
                node2_object, node2_found = find_checks(node2_name, node2_id)

                # Looking for owner =================================
                owner_object, owner_found = find_checks(owner_name, owner_id)

                # At this point - it is assumed that the user already ran objects and hierarchy and therefore objects are fixed
                edge_nodes_found = all([node1_found, node2_found])
                if not edge_nodes_found:
                    warn("Object not found...skipping line of Relations.", SyntaxWarning)
                    continue

                # found edge - modify
                if edge_found:
                    # Update Node 1, node2, name
                    edge_object.node1 = node1_object
                    edge_object.node2 = node2_object
                    edge_object.name = edge_name

                    # if no owner specified - assign to graph
                    if owner_found:  # owner specified
                        pass  # do nothing here
                    else:  # no owner specified
                        owner_object = self.graph  # specify graph as parent

                    # check ownership change
                    if edge_object.parent is not owner_object:  # if change in parent
                        # update ownership
                        # TODO: UPDATE EDGE OWNER
                        if owner_object:
                            # remove edge
                            edge_object.parent.remove_edge(edge_object)
                            # Add edge
                            new_edge = owner_object.add_edge(edge_object)
                        else:  # no  owner successfully specified
                            # add edge
                            assert isinstance(owner_object, Union([Group, Graph]))
                            new_edge = owner_object.add_edge(edge_object)

                    # Otherwise - just passing on edge for id recording
                    else:
                        new_edge = edge_object

                    # keep track of edges that were existing and still existing
                    edge_ids_after_mod.add(new_edge.id)

                # not found - create
                elif not edge_found:
                    edge_init_dict = dict()
                    edge_init_dict["node1"] = node1_object
                    edge_init_dict["node2"] = node2_object
                    edge_init_dict["name"] = edge_name
                    edge_init_dict = {key: value for (key, value) in edge_init_dict.items() if value is not None}
                    if not owner_object:
                        owner_object = self.graph
                    new_edge = owner_object.add_edge(**edge_init_dict)
                    edge_ids_after_mod.add(new_edge.id)

            # Deleting edges that have been deleted
            all_bulk_edge_ids = set(list(self.original_stats.all_edges.keys()))
            all_deleted_edge_ids = all_bulk_edge_ids.difference(edge_ids_after_mod)
            for del_edge_id in all_deleted_edge_ids:
                edge_obj: Edge = self.original_stats.all_edges[del_edge_id]
                parent = edge_obj.parent or self.graph
                parent.remove_edge(del_edge_id)

        elif self.type == "object_data":  # TODO: Implement management of deeper data - url, description, formatting
            raise NotImplementedError("This functionality is not yet implemented.")

        # Run Checks to avoid problems of manual data management
        self.graph.run_graph_rules()

    def give_user_chance_to_modify(self):
        """Open spreadsheet for user to modify data."""
        # kill spreadsheet process
        self.kill_spreadsheet()

        if show_guis:
            if app_platform == "Windows":
                os.startfile(self.TEMP_XLSX_WORKBOOK)
            elif app_platform == "Linux":
                command = f"xdg-open '{self.TEMP_XLSX_WORKBOOK}'"
                # output = execute_shell_command(command)
                subprocess.Popen(command, shell=True)

            user_response = msg.askokcancel(
                title="yEd Bulk Data Management - Async",
                message="To process changes, save and close workbook and press ok.",
            )

            if not user_response:
                print("User cancelled operation.")
                sys.exit()

    def bulk_data_management(self, type=None, graph=None, spreadsheet_data=None):
        """Process of converting to spreadsheet for manual operations, allows user to modify and then convert back to graph."""
        self.graph_to_spreadsheet_conversion(type=type, graph=graph)
        self.give_user_chance_to_modify()
        self.spreadsheet_to_graph_conversion(type=type, spreadsheet_data=spreadsheet_data)

        return self


class Graph:
    """Graph structure - carries yEd graph information"""

    def __init__(self, directed="directed", id="G"):
        self.directed = directed
        self.id = id

        self.nodes: dict[str, Node] = {}
        self.groups: dict[str, Group] = {}
        self.edges: dict[str, Edge] = {}
        self.combined_objects: dict[str, Union[(Node, Group)]] = {}

        self.custom_properties = []

        self.graphml: ET.Element

    # Addition of items ============================
    def add_node(self, node: Union[Node, str, None] = None, **kwargs) -> Node:
        """Adding node within Graph - accepts node object (simply assigns), or node name or none (to create new node)."""
        return add_node(self, node, **kwargs)

    def add_group(self, group: Union[Group, str, None] = None, **kwargs) -> Group:
        """Adding group to Graph"""
        return add_group(self, group, **kwargs)

    def add_edge(
        self,  # owner
        node1: Optional[Union[(Node, Group, str)]] = None,
        node2: Optional[Union[(Node, Group, str)]] = None,
        **kwargs,
    ) -> Edge:
        """Adding edge to Graph - for node1/node2 uses node / group objects or accepts names (creating new nodes under self) - or function can alternatively accept an instantiated Edge object."""

        # map args into kwargs in case of spreadsheet data management ops
        if node1:
            kwargs["node1"] = node1
        if node2:
            kwargs["node2"] = node2

        return add_edge(self, **kwargs)

    def define_custom_property(self, scope, name, property_type, default_value):
        """Adding custom properties to graph (which makes them available on the contained objects in yEd)"""
        if scope not in CUSTOM_PROPERTY_SCOPES:
            raise RuntimeWarning("Scope %s not recognised" % scope)
        if property_type not in CUSTOM_PROPERTY_TYPES:
            raise RuntimeWarning("Property Type %s not recognised" % property_type)
        if not isinstance(default_value, str):
            raise RuntimeWarning("default_value %s needs to be a string" % default_value)
        custom_property = CustomPropertyDefinition(scope, name, property_type, default_value)
        self.custom_properties.append(custom_property)
        if scope == "node":
            Node.set_custom_properties_defs(custom_property)
        elif scope == "edge":
            Edge.set_custom_properties_defs(custom_property)

    # Removal of items ==============================
    def remove_node(self, node: Union[Node, str]) -> None:
        """Remove/Delete a node from graph"""
        remove_node(self, node)

    def remove_group(self, group: Union[Group, str], **kwargs) -> None:
        """Removes a group from within current graph object (and same parent graph)."""
        remove_group(self, group, **kwargs)

    def remove_edge(self, edge: Union[Edge, str]) -> None:
        """Removing edge from graph - uses id."""
        remove_edge(self, edge)

    # TODO: ADD FUNCTIONALITY TO REMOVE / MODIFY CUSTOM PROPERTIES

    # Graph functionalities ===========================
    def construct_graphml(self) -> None:
        """Creating template graphml xml structure and then placing all graph items into it."""

        # Creating XML structure in Graphml format
        # xml = ET.Element("?xml", version="1.0", encoding="UTF-8", standalone="no")

        graphml = ET.Element("graphml", xmlns="http://graphml.graphdrawing.org/xmlns")
        graphml.set("xmlns:java", "http://www.yworks.com/xml/yfiles-common/1.0/java")
        graphml.set("xmlns:sys", "http://www.yworks.com/xml/yfiles-common/markup/primitives/2.0")
        graphml.set("xmlns:x", "http://www.yworks.com/xml/yfiles-common/markup/2.0")
        graphml.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
        graphml.set("xmlns:y", "http://www.yworks.com/xml/graphml")
        graphml.set("xmlns:yed", "http://www.yworks.com/xml/yed/3")
        graphml.set(
            "xsi:schemaLocation",
            "http://graphml.graphdrawing.org/xmlns http://www.yworks.com/xml/schema/graphml/1.1/ygraphml.xsd",
        )

        # Adding some implementation specific keys for identifying urls, descriptions
        node_key = ET.SubElement(graphml, "key", id="data_node")
        node_key.set("for", "node")
        node_key.set("yfiles.type", "nodegraphics")

        # Definition: url for Node
        node_key = ET.SubElement(graphml, "key", id="url_node")
        node_key.set("for", "node")
        node_key.set("attr.name", "url")
        node_key.set("attr.type", "string")

        # Definition: description for Node
        node_key = ET.SubElement(graphml, "key", id="description_node")
        node_key.set("for", "node")
        node_key.set("attr.name", "description")
        node_key.set("attr.type", "string")

        # Definition: url for Edge
        node_key = ET.SubElement(graphml, "key", id="url_edge")
        node_key.set("for", "edge")
        node_key.set("attr.name", "url")
        node_key.set("attr.type", "string")

        # Definition: description for Edge
        node_key = ET.SubElement(graphml, "key", id="description_edge")
        node_key.set("for", "edge")
        node_key.set("attr.name", "description")
        node_key.set("attr.type", "string")

        # Definition: Custom Properties for Nodes and Edges
        for prop in self.custom_properties:
            graphml.append(prop.convert_to_xml())

        edge_key = ET.SubElement(graphml, "key", id="data_edge")
        edge_key.set("for", "edge")
        edge_key.set("yfiles.type", "edgegraphics")

        # Graph node containing actual objects
        graph = ET.SubElement(graphml, "graph", edgedefault=self.directed, id=self.id)

        # Convert python graph objects into xml structure
        for node in self.nodes.values():
            graph.append(node.convert_to_xml())

        for group in self.groups.values():
            graph.append(group.convert_to_xml())

        for edge in self.edges.values():
            graph.append(edge.convert_to_xml())

        self.graphml = graphml

    def persist_graph(self, file=None, pretty_print=False, overwrite=False, vcs_version=False) -> File:
        """Convert graphml object->xml tree->graphml file.
        Temporary naming used if not given.
        """

        graph_file = File(file, extension=".graphml")

        if graph_file.file_exists and not overwrite:
            raise FileExistsError(f"File already exists: {graph_file.fullpath}")

        self.construct_graphml()

        if pretty_print:
            raw_str = ET.tostring(self.graphml)
            pretty_str = minidom.parseString(raw_str).toprettyxml()
            with open(graph_file.fullpath, "w") as f:
                f.write(pretty_str)
        else:
            tree = ET.ElementTree(self.graphml)
            tree.write(graph_file.fullpath)  # Uses internal method to XML Etree

        # Save simplified version for ease of vcs reviewability
        if vcs_version:
            # Extract file name / path information
            graph_file_name, _ = os.path.splitext(graph_file.fullpath)
            path = os.path.dirname(graph_file.fullpath)

            # Convert graph to spreadsheet =================
            new_spreadsheet = SpreadsheetManager()
            new_spreadsheet.graph_to_spreadsheet_conversion(type="relations", graph=self)

            # Convert to csv version =================

            # read into pandas spreadsheetfile
            xlsx = pd.ExcelFile(new_spreadsheet.TEMP_XLSX_WORKBOOK)

            # Iterate through each sheet in the spreadsheet file and save to csv
            for sheet_name in xlsx.sheet_names:
                # Read the sheet into a DataFrame
                df = pd.read_excel(xlsx, sheet_name)

                # Export the DataFrame to a CSV file
                csv_path = os.path.join(path, graph_file_name + "-" + str(sheet_name) + ".csv")
                df.to_csv(csv_path, index=False)

        # Update the file as existing or not
        graph_file.full_path_validate()

        print("persisting graph to file:", graph_file.fullpath)
        return graph_file

    def stringify_graph(self) -> str:
        """Returns Stringified version of graph in graphml format"""
        self.construct_graphml()
        # Py2/3 sigh.
        if sys.version_info.major < 3:
            return ET.tostring(self.graphml, encoding="UTF-8")
        else:
            return ET.tostring(self.graphml, encoding="UTF-8").decode()

    def from_existing_graph(self, file: str | File):
        """Parse GraphML xml of existing/stored graph file into python Graph structure."""

        id_existing_to_graph_obj = dict()

        # Manage file input ==============================
        if isinstance(file, File):
            graph_file = file
        else:
            graph_file = File(file, extension=".graphml")
        if not graph_file.file_exists:
            raise FileNotFoundError

        # Simplify input into string ==============================
        graph_str = xml_to_simple_string(graph_file.fullpath)

        # Begin XML parsing
        root = ET.fromstring(graph_str)

        # Extract off key information==============================
        all_keys = root.findall("key")
        key_dict = dict()
        for a_key in all_keys:
            sub_key_dict = dict()

            key_id = a_key.attrib.get("id")
            # sub_key_dict["label"] = a_key.attrib.get("for")
            sub_key_dict["attr"] = a_key.attrib.get("attr.name", None)
            # sub_key_dict["label"] = a_key.attrib.get("type")
            key_dict[key_id] = sub_key_dict

        # Get major graph node
        graph_root = root.find("graph")

        # get major graph info
        graph_dir = graph_root.get("edgedefault")
        graph_id = graph_root.get("id")

        # instantiate graph object
        new_graph = Graph(directed=graph_dir, id=graph_id)

        # Parse graph

        def is_group_node(node):
            return "foldertype" in node.attrib

        def process_node(parent, input_node):
            # Get sub nodes of this node (group or graph)
            current_level_nodes = input_node.findall("node")
            current_level_edges = input_node.findall("edge")

            for node in current_level_nodes:
                # normal nodes
                if not is_group_node(node):
                    node_init_dict = dict()

                    # <node id="n1">
                    existing_node_id = node.attrib.get("id", None)  # FIXME:

                    data_nodes = node.findall("data")
                    info_node = None
                    for data_node in data_nodes:
                        info_node = data_node.find("GenericNode") or data_node.find("ShapeNode")
                        if info_node is not None:
                            node_init_dict["node_type"] = info_node.tag

                            node_label = info_node.find("NodeLabel")
                            if node_label is not None:
                                node_init_dict["name"] = node_label.text

                                # TODO: PORT REST OF NODELABEL

                            # <Fill color="#FFCC00" transparent="false" />
                            fill = info_node.find("Fill")
                            if fill is not None:
                                node_init_dict["shape_fill"] = fill.get("color")
                                node_init_dict["transparent"] = fill.get("transparent")

                            # <BorderStyle color="#000000" type="line" width="1.0" />
                            border_style = info_node.find("BorderStyle")
                            if border_style is not None:
                                node_init_dict["border_color"] = border_style.get("color")
                                node_init_dict["border_type"] = border_style.get("type")
                                node_init_dict["border_width"] = border_style.get("width")

                            # <Shape type="rectangle" />
                            shape_sub = info_node.find("Shape")
                            if shape_sub is not None:
                                node_init_dict["shape"] = shape_sub.get("type")

                            uml = info_node.find("UML")
                            if uml is not None:
                                node_init_dict["shape"] = uml.get("AttributeLabel")
                            # TODO: THERE IS FURTHER DETAIL TO PARSE HERE under uml
                        else:
                            info = data_node.text
                            if info is not None:
                                info = re.sub(r"<!\[CDATA\[", "", info)  # unneeded schema
                                info = re.sub(r"\]\]>", "", info)  # unneeded schema

                                the_key = data_node.attrib.get("key")

                                info_type = key_dict[the_key]["attr"]
                                if info_type in ["url", "description"]:
                                    node_init_dict[info_type] = info
                    # Removing empty items
                    node_init_dict = {key: value for (key, value) in node_init_dict.items() if value is not None}
                    # create node
                    new_node = parent.add_node(**node_init_dict)
                    id_existing_to_graph_obj[existing_node_id] = new_node

                # group nodes
                # <node id="n2" yfiles.foldertype="group">
                elif is_group_node(node):
                    group_init_dict = dict()

                    # <node id="n1">
                    existing_group_id = node.attrib.get("id", None)

                    # Actual Group Data ===================================
                    data_nodes = node.findall("data")
                    for data_node in data_nodes:
                        proxy = data_node.find("ProxyAutoBoundsNode")
                        if proxy is not None:
                            realizer = proxy.find("Realizers")

                            group_nodes = realizer.findall("GroupNode")

                            for group_node in group_nodes:
                                geom_node = group_node.find("Geometry")
                                if geom_node is not None:
                                    group_init_dict["height"] = geom_node.attrib.get("height", None)
                                    group_init_dict["width"] = geom_node.attrib.get("width", None)
                                    group_init_dict["x"] = geom_node.attrib.get("x", None)
                                    group_init_dict["y"] = geom_node.attrib.get("y", None)

                                fill_node = group_node.find("Fill")
                                if fill_node is not None:
                                    group_init_dict["fill"] = fill_node.attrib.get("color", None)
                                    group_init_dict["transparent"] = fill_node.attrib.get("transparent", None)

                                borderstyle_node = group_node.find("BorderStyle")
                                if borderstyle_node is not None:
                                    group_init_dict["border_color"] = borderstyle_node.attrib.get("color", None)
                                    group_init_dict["border_type"] = borderstyle_node.attrib.get("type", None)
                                    group_init_dict["border_width"] = borderstyle_node.attrib.get("width", None)

                                nodelabel_node = group_node.find("NodeLabel")
                                if nodelabel_node is not None:
                                    group_init_dict["name"] = (
                                        nodelabel_node.text
                                    )  # TODO: SHOULD THIS JUST BE THE FIRST ONE?  IN OTHER WORDS - IS THERE MULTIPLE THINGS TO BE CAUGHT HERE?
                                    group_init_dict["font_family"] = nodelabel_node.attrib.get("fontFamily", None)
                                    group_init_dict["font_size"] = nodelabel_node.attrib.get("fontSize", None)
                                    group_init_dict["underlined_text"] = nodelabel_node.attrib.get(
                                        "underlinedText", None
                                    )
                                    group_init_dict["font_style"] = nodelabel_node.attrib.get("fontStyle", None)
                                    group_init_dict["label_alignment"] = nodelabel_node.attrib.get("alignment", None)

                                group_shape_node = group_node.find("Shape")
                                if group_shape_node is not None:
                                    group_init_dict["shape"] = group_shape_node.attrib.get("type", None)

                                group_state_node = group_node.find("State")
                                if group_state_node is not None:
                                    group_init_dict["closed"] = group_state_node.attrib.get("closed", None)
                                    # group_init_dict["aaa"] = group_state_node.attrib.get("closedHeight",None)
                                    # group_init_dict["aaaa"] = group_state_node.attrib.get("closedWidth",None)
                                    # group_init_dict["aaaa"] = group_state_node.attrib.get("innerGraphDisplayEnabled",None)

                                break

                        else:
                            info = data_node.text
                            if info is not None:
                                info = re.sub(r"<!\[CDATA\[", "", info)  # unneeded schema
                                info = re.sub(r"\]\]>", "", info)  # unneeded schema

                                the_key = data_node.attrib.get("key")

                                info_type = key_dict[the_key]["attr"]
                                if info_type in ["url", "description"]:
                                    group_init_dict[info_type] = info

                    # Group - Graph node
                    sub_graph_node = node.find("graph")

                    # Removing empty items
                    group_init_dict = {key: value for (key, value) in group_init_dict.items() if value is not None}

                    # Creating new group
                    new_group = parent.add_group(**group_init_dict)
                    id_existing_to_graph_obj[existing_group_id] = new_group

                    # Recursive processing
                    if sub_graph_node is not None:
                        process_node(parent=new_group, input_node=sub_graph_node)

                # unknown node type
                else:
                    raise NotImplementedError

            # edges then establish connections
            for edge_node in current_level_edges:
                edge_init_dict = dict()

                # <node id="n1">
                edge_id = edge_node.attrib.get("id", None)
                node1_id = edge_node.attrib.get("source", None)
                node2_id = edge_node.attrib.get("target", None)

                try:
                    edge_init_dict["node1"] = id_existing_to_graph_obj.get(node1_id)
                    edge_init_dict["node2"] = id_existing_to_graph_obj.get(node2_id)
                except Exception:  # TODO: MAKE MORE SPECIFIC
                    print(f"One of nodes of existing edge {edge_id} not found: {node1_id}, {node2_id} ")

                # FIXME: HOW TO MOVE FROM NODE IDS TO NODE OBJECTS - MOVE THROUGH GRAPHML FOR THE TEXT OF THAT OBJECT? - OR USE A DICTIONARY

                # <data key="d5">
                data_nodes = edge_node.findall("data")
                for data_node in data_nodes:
                    polylineedge = data_node.find("PolyLineEdge")

                    if polylineedge is not None:
                        # TODO: ADD POSITION MANAGEMENT
                        # path_node = polylineedge.find("Path")
                        # if path_node:
                        #   edge_init_dict["label"] = path_node.attrib.get("sx")
                        #   edge_init_dict["label"] = path_node.attrib.get("sy")
                        #   edge_init_dict["label"] = path_node.attrib.get("tx")
                        #   edge_init_dict["label"] = path_node.attrib.get("ty")

                        linestyle_node = polylineedge.find("LineStyle")
                        if linestyle_node is not None:
                            edge_init_dict["color"] = linestyle_node.attrib.get("color", None)
                            edge_init_dict["line_type"] = linestyle_node.attrib.get("type", None)
                            edge_init_dict["width"] = linestyle_node.attrib.get("width", None)

                        arrows_node = polylineedge.find("Arrows")
                        if arrows_node is not None:
                            edge_init_dict["arrowfoot"] = arrows_node.attrib.get("source", None)
                            edge_init_dict["arrowhead"] = arrows_node.attrib.get("target", None)

                        edgelabel_node = polylineedge.find("EdgeLabel")
                        if edgelabel_node is not None:
                            edge_init_dict["label"] = edgelabel_node.text
                            edge_init_dict["arrowfoot"] = edgelabel_node.attrib.get("source", None)
                            edge_init_dict["arrowhead"] = edgelabel_node.attrib.get("target", None)

                    else:
                        info = data_node.text
                        if info is not None:
                            info = re.sub(r"<!\[CDATA\[", "", info)  # unneeded schema
                            info = re.sub(r"\]\]>", "", info)  # unneeded schema

                            the_key = data_node.attrib.get("key")

                            info_type = key_dict[the_key]["attr"]
                            if info_type in ["url", "description"]:
                                edge_init_dict[info_type] = info

                # bendstyle_node = polylineedge.find("BendStyle")
                # edge_init_dict["smoothed"] = linestyle_node.attrib.get("smoothed") # TODO: ADD THIS

                # TODO:
                #   CUSTOM PROPERTIES

                # Removing empty items
                edge_init_dict = {key: value for (key, value) in edge_init_dict.items() if value is not None}
                parent.add_edge(**edge_init_dict)

        process_node(parent=new_graph, input_node=graph_root)

        return new_graph

    def manage_graph_data_in_spreadsheet(self, type: Optional[str] = None):
        """Port graph data into spreadsheet in several formats for easy and bulk creation and management.  Then ports back into python graph structure. Types: "obj_and_hierarchy", "object_data", "relations" """
        return SpreadsheetManager().bulk_data_management(graph=self, type=type)

    def gather_graph_stats(self) -> GraphStats:
        """Creating current Graph Stats for the current graph"""
        return GraphStats(graph=self)

    def run_graph_rules(self, correct: Optional[str] = None) -> None:
        """Check a few graph items that are most likely to fail following manual data management.  Correct them automatically or manually."""
        if correct is None:  #  ("auto", "manual")
            correct = "auto"

        stats = self.gather_graph_stats()

        def stranded_edges_check(self, graph_stats: GraphStats, correct: str) -> set[Edge]:
            """Check for edges with no longer valid nodes (these will prevent yEd from opening the file).  Correct them automatically or manually."""
            stranded_edges = set()
            for edge_id, edge in graph_stats.all_edges.items():
                node1_exist = edge.node1 in graph_stats.all_objects.values()
                node2_exist = edge.node2 in graph_stats.all_objects.values()
                at_least_one_edge = any([node1_exist, node2_exist])
                stranded_edge = not all([node1_exist, node2_exist])
                if stranded_edge:
                    stranded_edges.add(edge)

            if correct == "auto":
                for edge in stranded_edges:
                    edge.parent.remove_edge(edge)  # Any further postprocessing needed?

            elif correct == "manual":
                # spreadsheet - run relations and highlight edges with issues?
                raise NotImplementedError("Manual correction of stranded edges is not yet implemented.")

            # offer review or update edges
            return stranded_edges

        stranded_edges = stranded_edges_check(self, stats, correct)


# App related functions ===========================
def is_yed_findable():
    """Find yEd exe path locally"""
    path = which(PROGRAM_NAME) or None
    yed_found_bool = path is not None
    if not yed_found_bool:
        msg.showerror(
            title="Could not find yEd application.",
            message="Please install / add yEd to path.",
        )
    return yed_found_bool


def get_yed_process():
    """Return process object for yEd application, if there is one running."""
    process = None
    for process_iter in psutil.process_iter(["name"]):
        if process_iter.info["name"].startswith(PROGRAM_NAME):
            process = process_iter
            break
    return process


def get_yed_pid():
    """Return process id for yEd application or None if not running"""
    yed_pid = None

    if app_platform == "Windows":
        yed_process = get_yed_process()
        if yed_process:
            yed_pid = yed_process.pid
        return yed_pid
    elif app_platform == "Linux":
        command = "ps aux | grep yed.jar | grep -v grep | awk '{print $2}'"
        return execute_shell_command(command)


def is_yed_open() -> bool:
    """Check whether yEd is currently open - windows, linux, os, etc."""
    pid = get_yed_pid()
    return pid is not None


def execute_shell_command(command):
    try:
        # Run the command in the shell and capture the output
        result = subprocess.run(command, shell=True, capture_output=True, text=True, start_new_session=True)
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
    else:
        result = result.stdout.strip() if result.stdout else None
        return result


def start_subprocess(command):
    try:
        # Start the subprocess
        result = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return result
    except Exception as e:
        print(f"Unexpected error: {e}", file=sys.stderr)
    else:
        result = result.stdout.strip() if result.stdout else None
        return result
    return None


def open_yed_file(file: File, force=False):
    """Opens yed file - also will start yed if not open. Returns process or None."""

    def start():
        if app_platform == "Windows":
            os.startfile(file.fullpath)
        elif app_platform == "Linux":
            # command = [PROGRAM_NAME, file.fullpath]
            command = f"{PROGRAM_NAME} '{file.fullpath}'"
            # output = execute_shell_command(command)
            subprocess.Popen(command, shell=True, start_new_session=True)

        timer = 0
        time_limit = 5
        step = 0.5
        while not get_yed_pid() and timer < time_limit:
            sleep(step)
            timer += step
        if timer == time_limit:
            print("Failed attempt to start yEd application with file.")

    if not file.file_exists:
        return None

    if force:
        if show_guis:
            print("Act on yEd message box...")
            answer = msg.askokcancel(title="Force yEd App Close", message="Are you ok to force yEd closure?")
            if not answer:
                print("Exiting program")
                exit()

        kill_yed()
    start()

    if force:
        counter = 0
        start_attempts = 3
        while not get_yed_pid() and counter < start_attempts:
            start()
            start_attempts += 1
        if counter == start_attempts:
            print(f"Failed {counter} attempts at starting yEd application with file.")


def start_yed(wait=False):
    """Starts yEd program (if installed and on path and not already open)..

    Wait - if True, will wait for yEd to close before returning.
    """
    if is_yed_findable():
        if not is_yed_open():
            if wait is False:
                if os.name == "nt":  # Windows
                    subprocess.Popen("yEd", shell=True, creationflags=subprocess.CREATE_NEW_CONSOLE)
                else:  # macOS/Linux
                    subprocess.Popen("yEd", shell=True, start_new_session=True)

            else:
                print("Program waiting for yEd to be manually closed")
                subprocess.run(PROGRAM_NAME)
                # return None
    timer = 0
    time_limit = 8
    step = 0.5
    while not get_yed_pid() and timer < time_limit:
        sleep(step)
        timer += step
    if timer == time_limit:
        print("Problem with starting yEd application.")


def kill_yed() -> None:
    """Ends yEd program (if installed and on path and open)."""
    if app_platform == "Windows":
        yed_process = get_yed_process()
        if yed_process:
            yed_process.kill()
    elif app_platform == "Linux":
        pid = get_yed_pid()
        command = f"kill {pid}"
        execute_shell_command(command=command)

    timer = 0
    time_limit = 6
    step = 0.5
    while get_yed_pid() and timer < time_limit:
        sleep(step)
        timer += step
    if timer == time_limit:
        print("Problem with killing yEd application.")


# Utilities =======================================
def xml_to_simple_string(file_path) -> str:
    """Takes GraphML xml in string format and reduces complexity of the string for simpler parsing (without loss of any significant information).  Returns simplified string."""
    graph_str = ""
    try:
        with open(file_path, "r") as graph_file:
            graph_str = graph_file.read()

    except FileNotFoundError:
        print(f"Error, file not found: {file_path}")
        raise FileNotFoundError(f"Error, file not found: {file_path}")
    else:
        # Preprocessing of file for ease of parsing
        graph_str = graph_str.replace("\n", " ")  # line returns
        graph_str = graph_str.replace("\r", " ")  # line returns
        graph_str = graph_str.replace("\t", " ")  # tabs
        graph_str = re.sub("<graphml .*?>", "<graphml>", graph_str)  # unneeded schema
        graph_str = graph_str.replace("> <", "><")  # empty text
        graph_str = graph_str.replace("y:", "")  # unneeded namespace prefix
        graph_str = graph_str.replace("xml:", "")  # unneeded namespace prefix
        graph_str = graph_str.replace("yfiles.", "")  # unneeded namespace prefix
        graph_str = re.sub(" {1,}", " ", graph_str)  # reducing redundant spaces

    return graph_str


def generate_temp_uuid() -> str:
    """Temporary unique id for objects."""
    return str(randint(1, 1000000))


def assign_traceable_id(obj) -> None:
    """Creating unique traceable id for graph objects in similar format to yEd:
    n0, n1, ... for nodes and groups at a level
    e0, e1, ... for edges at a level
    n2::n2::n0 for nodes and groups at following level - full tracability
    n2::e0 for edges at following level - full tracability (lowest level ownership where linked)
    """
    parent_id_prefix = ""
    if isinstance(obj.parent, Group):
        parent_id_prefix = obj.parent.id + "::"

    if isinstance(obj, Node) or isinstance(obj, Group):
        # This object already logged under this owner - rename to order in list
        if obj in list(obj.parent.combined_objects.values()):
            obj_parent_index = list(obj.parent.combined_objects.values()).index(obj)
            obj.id = (
                parent_id_prefix + "n" + str(obj_parent_index)
            )  # FIXME: HAS TO BE THE INDEX OF THIS ITEM IN THE LIST
        # this item is new - appending to end of current dict - give new number for that level
        else:
            obj.id = parent_id_prefix + "n" + str(len(obj.parent.combined_objects))
    elif isinstance(obj, Edge):
        if obj in list(obj.parent.edges.values()):
            obj_parent_index = list(obj.parent.edges.values()).index(obj)
            obj.id = parent_id_prefix + "e" + str(obj_parent_index)
        else:
            obj.id = (
                parent_id_prefix + "e" + str(len(obj.parent.edges))
            )  # FIXME: HAS TO BE THE INDEX OF THIS ITEM IN THE LIST

    # print(obj.id)


def update_traceability(obj, owner, operation, heal=True) -> None:
    """Updating ownership of object based on parent."""

    if operation == "add":
        # Setting parent
        obj.parent = owner
        if isinstance(obj.parent, Group):
            obj.top_level_graph = obj.parent.top_level_graph
        else:
            obj.top_level_graph = obj.parent

        assign_traceable_id(obj)

        if isinstance(obj, Node):
            obj.parent.nodes[obj.id] = obj
            obj.parent.combined_objects[obj.id] = obj
        elif isinstance(obj, Group):
            obj.parent.groups[obj.id] = obj
            obj.parent.combined_objects[obj.id] = obj
        elif isinstance(obj, Edge):
            obj.parent.edges[obj.id] = obj

    if operation == "remove":
        if isinstance(obj, Node):
            del obj.parent.nodes[obj.id]
            del obj.parent.combined_objects[obj.id]

        elif isinstance(obj, Group):
            # reroute dependents ===================
            if heal:
                for node in obj.nodes.values():
                    node.parent = obj.parent  # reassign parent: node side
                    obj.parent.nodes[node.id] = node  # reassign parent: group side

                for edge in obj.edges.values():
                    edge.parent = obj.parent  # reassign parent: edge side
                    obj.parent.edges[edge.id] = edge  # reassign parent: group/graph side

                for group in obj.groups.values():
                    group.parent = obj.parent
                    obj.parent.groups[group.id] = group  # reassign parent: group side

            del obj.parent.groups[obj.id]
            del obj.parent.combined_objects[obj.id]

        elif isinstance(obj, Edge):
            del obj.parent.edges[obj.id]


# Reused graph functionality ============================
def add_node(owner, node, **kwargs) -> Node:
    """Adding node within Graph - accepts node object (simply assigns), or node name or none (to create new node)."""
    if isinstance(node, Node):
        # just update traceability - ownership
        pass
    if isinstance(node, str) or node is None:
        if node:
            kwargs["name"] = node
        node = Node(**kwargs)
    update_traceability(obj=node, owner=owner, operation="add")
    return node


def add_edge(owner: Union[(Graph, Group)], **kwargs) -> Edge:
    """Adding edge to graph -
    if node1/node2 names provided - creates nodes under the owner,
    otherwise, expects node / group objects at or under this owner level.
    Alternatively, can pass in an instantiated Edge object for ownership update."""

    edge = kwargs.get("edge", None)

    # If an edge is not passed in, create one
    if not edge:
        node1 = kwargs.get("node1", None)
        node2 = kwargs.get("node2", None)

        # Creating nodes if necessary from names - under owner
        if isinstance(node1, str):
            node1 = owner.add_node(node1)
            kwargs["node1"] = node1

        if isinstance(node2, str):
            node2 = owner.add_node(node2)
            kwargs["node2"] = node2

        # If not valid object, like None, should error out
        if not isinstance(node1, (Node, Group)):
            raise RuntimeWarning(f"Object {node1} doesn't exist")

        if not isinstance(node2, (Node, Group)):
            raise RuntimeWarning(f"Object {node2} doesn't exist")

        # http://graphml.graphdrawing.org/primer/graphml-primer.html#Nested
        # The edges between two nodes in a nested graph have to be declared in a graph,
        # which is an ancestor of both nodes in the hierarchy.

        if isinstance(owner, Group):
            if not (owner.is_ancestor(node1) and owner.is_ancestor(node2)):
                raise RuntimeWarning("Group %s is not ancestor of both %s and %s" % (owner.id, node1.id, node2.id))

        edge = Edge(**kwargs)

    # If an edge is passed in, continue on to traceability update
    else:
        pass

    update_traceability(obj=edge, owner=owner, operation="add")

    return edge


def add_group(owner, group, **kwargs) -> Group:
    """Adding group to graph"""
    if isinstance(group, Group):
        # just needs update to traceability - ownership
        pass
    if isinstance(group, str) or group is None:
        # if isinstance(owner, Group):
        #     top_level_graph = owner.top_level_graph
        # else:
        #     top_level_graph = owner
        if group:
            kwargs["name"] = group
        group = Group(**kwargs)
    heal: bool = kwargs.get("heal", True)
    update_traceability(obj=group, owner=owner, operation="add", heal=heal)
    return group


def remove_node(owner, node, **kwargs) -> None:
    """Remove/Delete a node - accepts node or node id"""
    if isinstance(node, Node):
        if node not in owner.nodes.values():
            raise RuntimeWarning(f"Node {node.id} doesn't exist")
    if isinstance(node, str):
        if node not in owner.nodes:
            raise RuntimeWarning(f"Node {node} doesn't exist")
        node = owner.nodes[node]
    update_traceability(obj=node, owner=owner, operation="remove")


def remove_group(owner, group, **kwargs) -> None:
    """Removes a group from within current object."""
    if isinstance(group, Group):
        if group not in owner.groups.values():
            raise RuntimeWarning(f"Group {group.id} doesn't exist")
    if isinstance(group, str):
        if group not in owner.groups:
            raise RuntimeWarning(f"Group {group} doesn't exist")
        group = owner.groups[group]

    update_traceability(obj=group, owner=owner, operation="remove")


def remove_edge(owner, edge, **kwargs) -> None:
    """Removing edge - uses id."""
    if isinstance(edge, Edge):
        if edge not in owner.edges.values():
            raise RuntimeWarning(f"Edge {edge.id} doesn't exist")
    if isinstance(edge, str):
        if edge not in owner.edges:
            raise RuntimeWarning(f"Edge {edge} doesn't exist")
        edge = owner.edges[edge]
    update_traceability(obj=edge, owner=owner, operation="remove")
