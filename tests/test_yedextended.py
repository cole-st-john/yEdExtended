"""Basic testing of yEdExtended functionality, including

* Building of yEd graph objects from scratch
* Reading of yEd graph files
* Management of yEd graph data in spreadsheet
* Opening and basic control of yEd application

"""

import io
import os
import xml.etree.ElementTree as xml

import openpyxl as pyxl
import pytest

import yedextended as yed
from yedextended import File, Graph, Node, SpreadsheetManager

# Triggers around testing completion
yed.testing = True
yed.show_guis = False
ci_system_check = os.environ.get("CI")
local_testing = yed.app_platform.startswith(("Windows", "Linux")) and not ci_system_check == "True"


class Test_File:
    """Testing File class"""

    def test_file_object_basics_1(self):
        """
        Given: File object
        When: given no basename or path
        Then: returns default graphml basename and working dir path"""

        test_file_obj = File()
        assert isinstance(test_file_obj, File)
        assert test_file_obj.basename == "temp.graphml"
        assert test_file_obj.window_search_name == "temp.graphml - yEd"
        assert test_file_obj.dir.lower() == os.getcwd().lower()

    def test_file_object_basics_2(self):
        """
        Given: File object
        When: given simple name w/o path
        Then: returns same basename and working dir path"""
        test_file_obj = File("abc.graphml")
        assert test_file_obj.basename == "abc.graphml"
        assert test_file_obj.window_search_name == "abc.graphml - yEd"
        assert test_file_obj.dir.lower() == os.getcwd().lower()

    def test_file_object_basics_3(self):
        """
        Given: File object
        When: given simple name w/ relative path (and cwd with that rel path)
        Then: returns same basename and working dir path"""
        test_file_obj = File("examples/abc.graphml")
        assert test_file_obj.basename == "abc.graphml"
        assert test_file_obj.window_search_name == "abc.graphml - yEd"
        assert test_file_obj.fullpath == os.path.join(os.getcwd(), "examples", "abc.graphml")
        assert os.path.exists(test_file_obj.dir) is True

    def test_file_object_basics_4(self):
        """
        Given: File object
        When: given simple name and bad path
        Then: returns same basename and diff path"""
        test_file_obj = File(r"c:/fakepath11/abc.graphml")
        assert test_file_obj.basename == "abc.graphml"
        assert test_file_obj.dir.lower() == os.getcwd().lower()

    def test_file_object_basics_5(self):
        """
        Given: File object
        When: given simple name and valid path
        Then: returns same basename and path"""
        test_file_obj = File(os.path.join(os.getcwd(), "abc.graphml"))
        assert test_file_obj.basename == "abc.graphml"
        assert test_file_obj.dir.lower() == os.getcwd().lower()


def test_graph_added_node_has_default_fill():
    """
    Given: new graph instance
    When: node added without any fill
    Then: shape fill should be expected default"""
    g = Graph()
    n1 = g.add_node("N1")
    assert "#FFCC00" == n1.shape_fill


def test_graph_added_node_keeps_custom_fill():
    """
    Given: new graph instance
    When: node added with fill
    Then: shape fill should be assigned"""
    g = Graph()
    n1 = g.add_node("N1", shape_fill="#99CC00")
    assert "#99CC00" == n1.shape_fill


def test_node_properties_after_nodes_and_edges_added():
    """
    Given: new graph instance
    When: nodes and edges added with shape and font style details
    Then: these details should carry these details"""
    g = Graph()

    node1 = g.add_node("foo", shape="ellipse")
    node2 = g.add_node("foo2", shape="roundrectangle", font_style="bolditalic")
    node3 = g.add_node("abc", shape="triangle", font_style="bold")
    edge1 = g.add_edge(node1, node2)

    assert node1.shape == "ellipse"
    assert node1.list_of_labels[0]._params["fontStyle"] == "plain"

    assert node2.shape == "roundrectangle"
    assert node2.list_of_labels[0]._params["fontStyle"] == "bolditalic"

    assert node3.shape == "triangle"
    assert node3.list_of_labels[0]._params["fontStyle"] == "bold"


def test_uml_node_properties_are_set():
    """
    Given: new graph instance
    When: UML node created with details (stereo, attrib, methods)
    Then: the node should be UML node and should carry these details"""
    g = Graph()

    expected_attributes = "int foo\nString bar"
    expected_methods = "foo()\nbar()"
    expected_stereotype = "abstract"

    abstract_node = g.add_node(
        "AbstractClass",
        node_type="UMLClassNode",
        UML={
            "stereotype": expected_stereotype,
            "attributes": expected_attributes,
            "methods": expected_methods,
        },
    )

    # UML Node
    graphml = g.stringify_graph()
    assertUmlNode(graphml, expected_stereotype, expected_attributes, expected_methods)

    # Carries UML info
    assert isinstance(abstract_node.UML, dict)
    assert abstract_node.UML["stereotype"] == expected_stereotype
    assert abstract_node.UML["attributes"] == expected_attributes
    assert abstract_node.UML["methods"] == expected_methods


def test_uml_stereotype_is_optional():
    """
    Given: new graph instance
    When: UML node created with details (attrib, methods)
    Then: the node should be UML node and should carry these details"""
    g = Graph()

    expected_attributes = "int foo\nString bar"
    expected_methods = "foo()\nbar()"

    class_node = g.add_node(
        "Class",
        node_type="UMLClassNode",
        UML={"attributes": expected_attributes, "methods": expected_methods},
    )

    graphml = g.stringify_graph()
    assertUmlNode(graphml, "", expected_attributes, expected_methods)

    assert g.nodes[class_node.id].UML["methods"] == expected_methods
    assert g.nodes[class_node.id].UML["attributes"] == expected_attributes


def assertUmlNode(graphml, expected_stereotype, expected_attributes, expected_methods):
    """Test helper function - examining uml contents"""
    doc = xml.fromstring(graphml)
    nsmap = {
        "g": "http://graphml.graphdrawing.org/xmlns",
        "y": "http://www.yworks.com/xml/graphml",
    }
    umlnode = doc.find("g:graph/g:node/g:data/y:UMLClassNode/y:UML", namespaces=nsmap)
    assert umlnode is not None

    attributes = umlnode.find("y:AttributeLabel", namespaces=nsmap)
    methods = umlnode.find("y:MethodLabel", namespaces=nsmap)

    assert attributes is not None
    assert methods is not None
    assert umlnode.attrib["stereotype"] == expected_stereotype
    assert attributes.text == expected_attributes
    assert methods.text == expected_methods


def test_numeric_node_ids():
    """
    Given: new graph instance
    When: adding nodes / edges
    Then: these items should carry the names as text and names, objects as well"""
    g = Graph()
    node1 = g.add_node("Node1")
    node2 = g.add_node("Node2")
    edge1 = g.add_edge(node1, node2)

    assert node1.list_of_labels[0]._text == "Node1"
    assert node2.list_of_labels[0]._text == "Node2"
    assert node1.name == "Node1"
    assert node2.name == "Node2"

    assert isinstance(edge1.node1, Node)
    assert isinstance(g.edges[edge1.id].node2, Node)


def test_multiple_edges():
    """
    Given: new graph instance
    When: adding labels to node, edge, graph
    Then: these items should carry the names as text and names, objects as well"""
    g = Graph()
    node1 = g.add_node("a").add_label("a2")
    edge1 = g.add_edge("b", "c", name="d").add_label("d2")
    group1 = g.add_group("e")
    edge2 = group1.add_edge("f", "g", name="h").add_label("h2")

    assert node1.id == "n0"
    assert node1.name == "a"
    assert node1.list_of_labels[0]._text == "a"
    assert node1.list_of_labels[1]._text == "a2"

    assert group1.id == "n3"
    assert group1.name == "e"

    assert edge1.id == "e0"
    assert edge1.name == "d"
    assert edge1.list_of_labels[0]._text == "d"
    assert edge1.list_of_labels[1]._text == "d2"

    assert edge2.id == "n3::e0"
    assert edge2.name == "h"
    assert edge2.list_of_labels[0]._text == "h"
    assert edge2.list_of_labels[1]._text == "h2"


def test_node_already_there_check_1():
    """
    Given: new graph instance
    When: adding dup node names
    Then: no error is raised"""
    g = Graph()
    node1 = g.add_node("a")
    node2 = g.add_node("a")
    assert node1 and node2
    assert node1 is not node2


def test_node_already_there_check_2():
    """
    Given: new graph instance
    When: adding dup node obj
    Then: no error is raised"""
    g = Graph()
    node1 = g.add_node("a")
    node2 = g.add_node(node1)
    assert node1 and node2
    assert node1 is node2


def test_node_already_there_check_3():
    """
    Given: new graph instance
    When: adding edge with nodes that dont exist
    Then: error raised"""
    g = Graph()
    with pytest.raises(NameError):  # FIXME: ADD IN INIT
        g.add_edge(a, b)


def test_node_already_there_check_4():
    """
    Given: new graph instance
    When: adding group with names that exist
    Then: error raised"""
    g = Graph()
    g1 = g.add_group("g1")
    g2 = g.add_group("g1")
    assert g1 is not g2


def test_nested_graph_edges():
    """
    Given: new graph instance
    When: numerous items added
    Then: expected parent-age recorded as expected"""
    g = Graph()

    a = g.add_node("a")
    b = g.add_node("b")
    edge1 = g.add_edge(a, b)
    assert len(g.edges) == 1

    g1 = g.add_group("g1")
    g1n1 = g1.add_node("g1n1")
    g1n2 = g1.add_node("g1n2")
    g2 = g1.add_group("g2")
    g2n1 = g2.add_node("g2n1")
    g2n2 = g2.add_node("g2n2")
    g3 = g1.add_group("g3")
    g3n1 = g3.add_node("g3n1")
    g3n2 = g3.add_node("g3n2")

    g1.add_edge(g1n1, g1n2)
    assert len(g.edges) == 1
    assert len(g1.edges) == 1

    g2.add_edge(g2n2, g2n2)  # cycle - not a typo
    assert len(g2.edges) == 1

    g3.add_edge(g3.add_node("c"), g3.add_node("d"))
    assert len(g3.edges) == 1

    g.add_edge(g2n1, g2n2)
    g1.add_edge(g2n1, g2n2)
    g2.add_edge(g2n1, g2n2)
    assert len(g.edges) == 2


def test_xml_to_simple_string_1():
    """
    Given: yEd utility
    When: valid file with items needing simplification
    Then: returns simplified string"""
    test_string = yed.xml_to_simple_string("examples/yed_created_edges.graphml")
    assert test_string.find("\n") == -1


def test_xml_to_simple_string_2():
    """
    Given: yEd utility
    When: invalid file
    Then: returns exception"""

    with pytest.raises(FileNotFoundError):
        yed.xml_to_simple_string("not_existing_file")


class Test_GraphStats:
    """Basic tests of graph statisics"""

    def test_graph_stats_basic_1(self):
        """
        Given: existing graph of known stats
        When: gathering statistics
        Then: statistics as expected"""
        test_graph = Graph().from_existing_graph("examples/yed_created_edges.graphml")
        results_stats = test_graph.gather_graph_stats()

        # Then: we can assume the following stats
        assert len(results_stats.all_edges) == 3
        assert len(results_stats.all_groups) == 1
        assert len(results_stats.all_nodes) == 4
        assert len(results_stats.all_graph_items) == 8
        assert len(results_stats.all_objects) == 5

    def test_graph_stats_basic_2(self):
        """
        Given: empty graph
        When: gathering statistics
        Then: statistics as expected"""

        test_graph = Graph().from_existing_graph("examples/yed_created_empty_graph.graphml")

        # When: taking the stats
        results_stats = test_graph.gather_graph_stats()

        # Then: we can assume the following stats
        assert len(results_stats.all_edges) == 0
        assert len(results_stats.all_groups) == 0
        assert len(results_stats.all_nodes) == 0
        assert len(results_stats.all_graph_items) == 0
        assert len(results_stats.all_objects) == 0


def test_round_trip():
    """
    Given: basic graph
    When: persist_graph used, then from_existing_graph used ("round-robin" translation)
    Then: resulting graph matches the in-memory object"""

    FILE = "roundtrip.graphml"
    # cleanup
    if os.path.exists(FILE):
        os.remove(FILE)

    # basic graph
    graph = Graph()
    a = graph.add_node("a")
    b = graph.add_node("b")
    graph.add_edge(a, b)
    graph_file = graph.persist_graph(FILE)

    # grabbing stored for comparison
    graph_after = Graph().from_existing_graph(graph_file)

    # comparison
    assert graph.stringify_graph() == graph_after.stringify_graph()

    # making unequal
    graph.add_node("c")

    # retesting
    assert graph.stringify_graph() != graph_after.stringify_graph()

    # cleanup
    if os.path.exists(FILE):
        os.remove(FILE)


def test_custom_property_assignment():
    """
    Given: graph with some custom properties assigned
    When: testing for the properties
    Then: all as expected (findable and correct)"""
    graph1 = Graph()

    graph1.define_custom_property("node", "Population", "int", "0")
    graph1.define_custom_property("edge", "Population", "int", "1")
    group1 = graph1.add_group("group1", custom_properties={"Population": "2"})
    node1 = graph1.add_node("a")
    edge1 = graph1.add_edge("group1", "a")
    node2 = group1.add_node("b")
    node3 = group1.add_node("c", custom_properties={"Population": "3"})

    assert node1.Population == "0", "Property not as expected"
    assert node2.Population == "0", "Property not as expected"
    assert node3.Population == "3", "Property not as expected"
    assert edge1.Population == "1", "Property not as expected"
    assert group1.Population == "2", "Property not as expected"

    assert graph1.stringify_graph().find("Population") != -1, "Property not found in graphml"


def test_persist_graph_1():
    """
    Given: simple graph
    When: persist graph used
    Then: file exists"""

    file1 = "abcd.graphml"

    # Ensure cleaned up
    if os.path.exists(file1):
        os.remove(file1)

    graph1 = Graph()
    graph1.add_node("a")
    graph1.add_node("b")

    # Test that a file is created/stored
    graph1.persist_graph(file1)
    assert os.path.exists(file1) is True

    # Ensure cleaned up
    if os.path.exists(file1):
        os.remove(file1)


def test_persist_graph_2():
    """
    Given: simple graph
    When: persist graph used
    Then: errors only occur if duplicate but no overwrite"""
    file1 = "abcd.graphml"

    # Ensure cleaned up
    if os.path.exists(file1):
        os.remove(file1)

    graph1 = Graph()
    graph1.add_node("a")
    graph1.add_node("b")

    # Test redundant file error - if same filename used
    graph1.persist_graph(file1)
    with pytest.raises(FileExistsError):
        Graph().persist_graph(file1)

    # Test no error if overwrite is True
    Graph().persist_graph(file1, overwrite=True)

    # Ensure cleaned up
    if os.path.exists(file1):
        os.remove(file1)


def test_persist_graph_3():
    """
    Given: simple graph
    When: persist graph used with or without pretty print
    Then: pretty print evident through line returns/tab"""

    file1 = "abcd.graphml"
    file2 = "abcde.graphml"

    # Ensure cleaned up
    if os.path.exists(file1):
        os.remove(file1)
    if os.path.exists(file2):
        os.remove(file2)

    graph1 = Graph()
    graph1.add_node("a")
    graph1.add_node("b")

    # Test pretty print not same contents as non-pretty print
    graph1.persist_graph(file1)
    graph1.persist_graph(file2, pretty_print=True)
    file1_handle = open(file1, "r")
    file2_handle = open(file2, "r")
    file1_contents = file1_handle.read()
    file2_contents = file2_handle.read()
    file1_handle.close()
    file2_handle.close()
    assert file1_contents != file2_contents
    assert file2_contents.count("\n") > file1_contents.count("\n")
    assert file2_contents.count("\t") > file1_contents.count("\t")

    # Ensure cleaned up
    if os.path.exists(file1):
        os.remove(file1)
    if os.path.exists(file2):
        os.remove(file2)


def test_from_existing_graph_1():
    """
    Given: use of from_existing_graph
    When: graph file is existing
    Then: return graph object expected"""
    assert isinstance(Graph().from_existing_graph("examples/yed_created_edges.graphml"), Graph)


def test_from_existing_graph_2():
    """
    Given: use of from_existing_graph
    When: graph file is not existing
    Then: return error"""
    with pytest.raises(FileNotFoundError):
        Graph().from_existing_graph("not_existing_file")


def test_from_existing_graph_3():
    """
    Given: use of from_existing_graph with graph file is existing and simple graph
    When: graph stats used
    Then: expected information returned"""
    test_graph = Graph().from_existing_graph("examples/yed_created_edges.graphml")
    test_graph_stats = test_graph.gather_graph_stats()
    assert test_graph_stats.all_nodes["n0"].url is not None
    assert len(test_graph_stats.all_graph_items) == 8


def test_create_graph_with_url_description():
    """
    Given: simple graph
    When: when edges, nodes, groups are added with url, description
    Then: the info is findable in graph object"""
    FILE1 = "test.graphml"

    graph1 = Graph()

    url = "http://www.google.com"
    description = "This is an node with a URL and description"
    node1 = graph1.add_node(
        "a",
        url=url,
        description=description,
    )
    graph1.add_node("b")
    assert node1.url == url
    assert node1.description == description

    url = "http://www.google.com"
    description = "This is an edge with a URL and description"
    edge1 = graph1.add_edge(
        "a",
        "b",
        url=url,
        description=description,
    )
    assert edge1.url == url
    assert edge1.description == description

    url = ("http://www.google.com",)
    description = "This is a group with a URL and description"
    group1 = graph1.add_group(
        "group1",
        url=url,
        description=description,
    )
    assert group1.url == url
    assert group1.description == description


def test_create_edge_w_o_nodes():
    """
    Given: empty graph
    When: edge is added with node names
    Then: nodes are created with these names at the expected level (edge owner)"""
    graph1 = Graph()
    edge = graph1.add_edge("a", "b")

    node_a = edge.node1
    assert isinstance(node_a, Node)
    assert graph1.nodes[node_a.id] is not None


def test_removes():
    """
    Given: simple graph instance
    When: items are removed (nodes / groups / edges)
    Then: the expected changes occur - ownership links are updated"""
    # Creating simple graph ================================
    graph1 = Graph()

    # first level nodes
    a = graph1.add_node("a")
    b = graph1.add_node("b")
    c = graph1.add_node("c")

    # first level edges
    edge1 = graph1.add_edge(a, b)
    edge2 = graph1.add_edge(a, c)

    # first level group
    group1 = graph1.add_group("group1")

    # second level items
    group1.add_node("d")
    e = group1.add_node("e")
    f = group1.add_node("f")
    edge3 = group1.add_edge(e, f)

    # third level items
    group1_1 = group1.add_group("group1_1")

    # gather graph stats =========================
    stats1 = graph1.gather_graph_stats()

    # Examining graph stats ==========================
    assert f in stats1.all_nodes.values()
    assert e in group1.nodes.values()
    assert a in stats1.all_nodes.values()

    assert group1_1 in stats1.all_groups.values()
    assert group1_1 in group1.groups.values()
    assert group1 in graph1.groups.values()

    assert edge1 in stats1.all_edges.values()
    assert edge2 in stats1.all_edges.values()
    assert edge3 in stats1.all_edges.values()
    assert edge3 in group1.edges.values()

    graph1.remove_node(b)
    graph1.remove_group(group1)

    stats2 = graph1.gather_graph_stats()

    assert f in stats2.all_nodes.values()
    assert e in group1.nodes.values()
    assert a in stats1.all_nodes.values()
    assert group1_1 in stats2.all_groups.values()
    assert group1_1 in graph1.groups.values()
    assert group1 not in graph1.groups.values()

    assert edge1 in stats2.all_edges.values()
    assert edge2 in stats2.all_edges.values()
    assert edge3 in stats2.all_edges.values()
    assert edge3 in graph1.edges.values()

    graph1.run_graph_rules()

    stats3 = graph1.gather_graph_stats()

    # check for common errors
    with pytest.raises(RuntimeWarning):
        graph1.remove_node("na")

    with pytest.raises(RuntimeWarning):
        graph1.remove_edge("na")

    with pytest.raises(RuntimeWarning):
        graph1.remove_group("group_na")


@pytest.mark.skipif(
    local_testing is not True,
    reason="Tests not suitable for CI / Non-windows environments at this time",
)
class Test_Yed_App_Functions:
    """Testing yed app related functionalities - assumes yEd installed (likely not true on test server)"""

    def test_start_yed(self):
        """
        Given:
        When: starting yed
        Then: returns valid process object"""
        yed.kill_yed()
        yed.start_yed()
        assert yed.get_yed_pid() is not None, "Expected a PID but got none"
        yed.kill_yed()

    def test_basic_findable_and_open_file(self):
        """
        Given: yed is installed
        When: is_yed_findable / open_with_yed used with basic file
        Then: findable and returns valid process object"""

        # Initialize the test File
        test_graph = File("examples/yed_created_edges.graphml")

        # Assert that YED is findable
        assert yed.is_yed_findable() is True

        # Open the File asynchronously and await its completion - requires manual closure
        test_graph.open_with_yed()

        # Assert that the result is not None (after user closed yEd during test)
        assert yed.get_yed_pid() is not None, "Expected a PID, but got None"

        # Kill the YED process
        yed.kill_yed()  # redundant

    def test_file_object_app(self):
        """
        Given: yEd is installed
        When: triggering open_with_yed with valid relative file path
        Then: file should be opened"""
        test_file_obj = File("examples/test.graphml")
        pid = test_file_obj.open_with_yed()
        assert pid is not None, "Expected a PID, but got None"

        # clean up
        yed.kill_yed()

    def test_yed_kill_then_check1(self):
        """
        Given: yEd is installed
        When: yEd started and triggering yed kill
        Then: yEd closed / no longer running"""

        # kill any yEd
        yed.kill_yed()
        # check for yed running
        assert yed.is_yed_open() is False

    def test_yed_open_then_check1(self):
        """
        Given: yEd is installed
        When: yEd started and triggering yed kill
        Then: yEd closed / no longer running"""
        # start yEd
        yed.start_yed()

        assert yed.is_yed_open() is True

    def test_yed_kill_then_check2(self):
        """
        Given: yEd is installed
        When: yEd started and triggering yed kill
        Then: yEd closed / no longer running"""

        # start yEd
        yed.start_yed()
        yed.kill_yed()

        assert yed.is_yed_open() is False

    def test_yed_kill_then_check3(self):
        yed.start_yed()
        assert yed.get_yed_pid() is not None

    def test_yed_start_w_manual_close_wait(self):
        """
        Given: start_yed is used
        When: when wait is true
        Then: return process should be none (it is already closed)"""
        yed.kill_yed()
        yed.start_yed(wait=True)
        assert yed.get_yed_pid() is None, "Expected yEd PID, but got None"

    def test_open_yed_file(self):
        """
        Given: use of open_yed_file
        When: handed real graph file
        Then: return process object expected"""
        graph_file = File("examples/yed_created_edges.graphml")

        yed.open_yed_file(graph_file)
        assert yed.get_yed_pid() is not None, "Expected an active PID, but got None"
        yed.kill_yed()

        yed.open_yed_file(graph_file, force=True)
        assert yed.get_yed_pid() is not None, "Expected an active PID, but got None"
        yed.kill_yed()

        graph_file = File("not_real_file.graphml")
        yed.open_yed_file(graph_file)
        assert yed.get_yed_pid() is None, "Expected no yEd process should be open - instead one is open."


@pytest.mark.skipif(
    local_testing is not True,
    reason="Tests not suitable for CI / Non-windows environments at this time",
)
class Test_Spreadsheet_Related_Functionalities:
    """Test all spreadsheet based functionality - assumes spreadsheet app"""

    def test_init(self):
        """
        Given: empty graph
        When: graph_to_spreadsheet is used
        Then: it is expected that a template file is created and empty (except for headers)
        """
        spreadsheet = SpreadsheetManager()
        graph = Graph()
        spreadsheet.graph_to_spreadsheet_conversion(graph=graph)
        assert spreadsheet is not None
        assert os.path.isfile(spreadsheet.TEMP_XLSX_WORKBOOK) is True, "Expected template created"

    def test_graph_to_spreadsheet_conversion_obj(self):
        """
        Given: simple graph instance from existing graphml
        When: graph_to_spreadsheet conversion is used
        Then: the object output should be as expected"""

        graph1 = Graph().from_existing_graph("examples/yed_created_edges.graphml")
        # test conversion to spreadsheet
        spreadsheet1 = SpreadsheetManager()
        spreadsheet1.graph_to_spreadsheet_conversion(graph=graph1)

        in_mem_file1 = None
        in_mem_file2 = None
        with open(spreadsheet1.TEMP_XLSX_WORKBOOK, "rb") as f:
            in_mem_file1 = io.BytesIO(f.read())

        with open("examples/yed_test_to_spreadsheet1.xlsx", "rb") as f:
            in_mem_file2 = io.BytesIO(f.read())

        current = pyxl.load_workbook(in_mem_file1).active
        reference = pyxl.load_workbook(in_mem_file2).active
        current_data = self.get_filtered_sheet_values(current)
        reference_data = self.get_filtered_sheet_values(reference)
        assert current_data == reference_data

        if os.path.exists(spreadsheet1.TEMP_XLSX_WORKBOOK):
            os.remove(spreadsheet1.TEMP_XLSX_WORKBOOK)

    def test_graph_to_spreadsheet_conversion_rel(self):
        """
        Given: simple graph instance from existing graphml
        When: graph_to_spreadsheet conversion is used
        Then: the relation output should be as expected"""

        graph1 = Graph().from_existing_graph("examples/yed_created_edges.graphml")
        # test conversion to spreadsheet
        spreadsheet1 = SpreadsheetManager()
        spreadsheet1.graph_to_spreadsheet_conversion(graph=graph1, type="relations")

        in_mem_file1 = None
        in_mem_file2 = None
        with open(spreadsheet1.TEMP_XLSX_WORKBOOK, "rb") as f:
            in_mem_file1 = io.BytesIO(f.read())

        with open("examples/yed_test_to_spreadsheet2.xlsx", "rb") as f:
            in_mem_file2 = io.BytesIO(f.read())

        current = pyxl.load_workbook(in_mem_file1)["Relations"]
        reference = pyxl.load_workbook(in_mem_file2)["Relations"]
        current_data = self.get_filtered_sheet_values(current)
        reference_data = self.get_filtered_sheet_values(reference)
        assert current_data == reference_data

        if os.path.exists(spreadsheet1.TEMP_XLSX_WORKBOOK):
            os.remove(spreadsheet1.TEMP_XLSX_WORKBOOK)

    def test_bulk_data_management(self):
        """
        Given: simple graph instance (hardcoded) with multiple layers
        When: manage_graph_data  is used but without manual manipulation
        Then: it is expected that no changes are made to the graph itself"""
        graph1 = Graph()

        # first level nodes
        graph1.add_node("a")
        graph1.add_node("b")
        graph1.add_node("c")

        # first level group
        group1 = graph1.add_group("group1")

        # second level items
        group1.add_node("d")
        group1_1 = group1.add_group("group1_1")
        group1_1.add_node("e")

        # shallow copy the stringified graph
        before_string = graph1.stringify_graph()[:]

        # doing nothing in spreadsheet should lead to the same graph being built back up (when talking about simple obj model)
        spreadsheet_obj = graph1.manage_graph_data_in_spreadsheet()

        after_string = graph1.stringify_graph()

        assert before_string is not None
        assert after_string is not None
        assert before_string == after_string

        if os.path.exists(spreadsheet_obj.TEMP_XLSX_WORKBOOK):
            os.remove(spreadsheet_obj.TEMP_XLSX_WORKBOOK)

    def test_spreadsheet_to_graph(self):  # FIXME:
        """
        Given: simple graph instance (hardcoded)
        When: spreadsheet_to_graph is used as expected
        Then: the resulting graph from the template spreadsheet (produced with the graph_to_spreadsheet)
        """
        # Create Graph
        graph1 = yed.Graph()
        n2 = graph1.add_node("Ivrea")
        n1 = graph1.add_node("Turin")
        group1 = graph1.add_group("Northern Italy")
        n3 = group1.add_node("Savona")
        n4 = group1.add_node("Brescia")
        graph1.add_edge(n1, n4)
        graph1.add_edge(n1, n2)
        graph1.add_edge(n1, n3)

        # test conversion to spreadsheet
        spreadsheet1 = SpreadsheetManager()
        data = "examples/yed_test_to_spreadsheet3.xlsx"
        spreadsheet1.spreadsheet_to_graph_conversion(type="obj_and_hierarchy", spreadsheet_data=data)
        spreadsheet1.spreadsheet_to_graph_conversion(type="relations", spreadsheet_data=data)
        reference_string = graph1.stringify_graph()
        return_string = spreadsheet1.graph.stringify_graph()
        assert reference_string == return_string

    def test_spreadsheet_round_trip_change_1(self):
        """
        Given: simple graph instance from existing graphml
        When: round trip  conversion is mocked - but with changes to existing
        Then: the object output should be as expected"""

        # n0 (ivrea): node -> group
        # n1 (turin): ownership graph -> ivrea
        # n2 (northern ..): group -> node
        #  (savona ..): ownership northern -> graph
        #  (brescia ..): ownership northern -> graph

        # Get graph
        spreadsheet1 = SpreadsheetManager()
        data1 = "examples/yed_test_to_spreadsheet4.xlsx"
        spreadsheet1.spreadsheet_to_graph_conversion(type="obj_and_hierarchy", spreadsheet_data=data1)
        spreadsheet1.spreadsheet_to_graph_conversion(type="relations", spreadsheet_data=data1)
        print("")
        data2 = "examples/yed_test_to_spreadsheet4_mod1.xlsx"
        spreadsheet1.spreadsheet_to_graph_conversion(type="obj_and_hierarchy", spreadsheet_data=data2)
        # spreadsheet1.spreadsheet_to_graph_conversion(type="relations", spreadsheet_data=data2)
        # print("h")

    def test_spreadsheet_round_trip_change_2(self):
        """
        Given: simple graph instance from existing graphml
        When: some changes are made to edges
        Then: the object output should be as expected"""

        # Get graph
        spreadsheet1 = SpreadsheetManager()
        data1 = "examples/yed_test_to_spreadsheet4.xlsx"
        spreadsheet1.spreadsheet_to_graph_conversion(type="obj_and_hierarchy", spreadsheet_data=data1)
        spreadsheet1.spreadsheet_to_graph_conversion(type="relations", spreadsheet_data=data1)

        data2 = "examples/yed_test_to_spreadsheet4_mod2.xlsx"
        spreadsheet1.spreadsheet_to_graph_conversion(type="relations", spreadsheet_data=data2)

    def get_filtered_sheet_values(self, sheet):
        """test helper function - Filter empty columns from spreadsheet data for test comparison"""
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Filter out empty columns
            filtered_row = [cell for cell in row if cell is not None]
            if filtered_row:  # Only add non-empty rows
                data.append(filtered_row)
        return data


def test_edge_with_custom_props_1():
    """
    Given: simple graph
    When: adding edge with custom props
    Then: the object output should be as expected"""

    url = "www.google.com"
    description = "test of edge props"

    graph = Graph()

    graph.define_custom_property("edge", "Distance", "int", "0")
    graph.define_custom_property("edge", "Availability", "double", "100.0")
    graph.define_custom_property("edge", "Toll Free", "boolean", "true")
    graph.define_custom_property("edge", "Year of build", "string", "")

    edge1 = graph.add_edge(
        "Node1",
        "Node2",
        name="Edge1",
        custom_properties={
            "Year of build": "1974",
            "Distance": "356",
            "Toll Free": "false",
            "Availability": "85.7",
        },
        url=url,
        description=description,
    )

    file = "temp.graphml"
    if os.path.exists(file):
        os.remove(file)

    graph.persist_graph()

    assert edge1.url == url

    if os.path.exists(file):
        os.remove(file)


def test_edge_with_custom_props_2():
    """
    Given: simple graph
    When: adding edge with custom props
    Then: the object output should be as expected"""

    url = "www.google.com"
    description = "test of edge props"

    graph = Graph()

    graph.define_custom_property("node", "Population", "int", "0")
    graph.define_custom_property("node", "Unemployment", "double", "0.0")
    graph.define_custom_property("node", "Environmental Engagements", "boolean", "false")
    graph.define_custom_property("node", "Mayor", "string", "")
    graph.define_custom_property("node", "Country", "string", "")

    node1 = graph.add_node(
        "Random City",
        custom_properties={
            "Population": "13000",
            "Unemployment": "13.7",
            "Environmental Engagements": "true",
            "Mayor": "Genarro",
        },
        url=url,
        description=description,
    )

    file = "temp.graphml"
    if os.path.exists(file):
        os.remove(file)

    graph.persist_graph()

    assert node1.url == url
    assert node1.description == description

    if os.path.exists(file):
        os.remove(file)


def test_node_geoms():
    """
    Given: simple graph
    When: adding node with geom traits
    Then: the object output should be as expected"""

    file = "temp.graphml"
    if os.path.exists(file):
        os.remove(file)

    graph = Graph()
    graph.add_node("Node1", height="50", width="100", x="100", y="100")
    graph.persist_graph()

    if os.path.exists(file):
        os.remove(file)
